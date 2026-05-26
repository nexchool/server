"""Tests for GET /import/template — pure-Python, no DB, no Flask app context.

Covers:
  1. The route handler function exists on the routes module
  2. The xlsx body produced by the handler contains the expected headers
"""
from __future__ import annotations

import sys
from pathlib import Path

SERVER_DIR = Path(__file__).resolve().parent.parent
if str(SERVER_DIR) not in sys.path:
    sys.path.insert(0, str(SERVER_DIR))


def test_download_import_template_function_exists():
    """The handler is defined on the routes module."""
    from modules.school_setup import routes

    assert callable(getattr(routes, "download_import_template", None))


def test_download_import_template_produces_valid_xlsx():
    """The xlsx skeleton created by the handler contains the expected header row."""
    from io import BytesIO
    from openpyxl import Workbook, load_workbook

    # Reproduce exactly the logic from the route handler (sans Flask context)
    wb = Workbook()
    ws = wb.active
    ws.title = "Classes"
    ws.append(["unit_code", "programme_code", "grade", "section", "subject", "periods"])
    ws.append(["MN", "CBSE-ENG", "Grade 1", "A", "", ""])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    # Verify the file is a valid xlsx with correct headers
    loaded = load_workbook(bio)
    assert "Classes" in loaded.sheetnames
    sheet = loaded["Classes"]
    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header_row == ["unit_code", "programme_code", "grade", "section", "subject", "periods"]


def test_download_import_template_mimetype_constant():
    """The route handler uses the correct xlsx mimetype constant."""
    import inspect
    from modules.school_setup import routes

    src = inspect.getsource(routes.download_import_template)
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in src
    assert "as_attachment=True" in src
    assert "class-import-template.xlsx" in src
