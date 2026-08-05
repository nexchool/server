"""Academic calendar module — wizard state, working-day math, publish flow.

DB-backed tests using the savepoint-rollback fixtures from conftest.py.
Services read g.tenant_id, so calls run inside test_request_context.
"""
from __future__ import annotations

import sys
import uuid
from datetime import date
from pathlib import Path

import pytest
from flask import g

SERVER_DIR = Path(__file__).resolve().parent.parent
if str(SERVER_DIR) not in sys.path:
    sys.path.insert(0, str(SERVER_DIR))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_year(db_session, tenant, name="2026-2027",
               start="2026-06-01", end="2026-06-30"):
    from modules.academics.academic_year.models import AcademicYear

    year = AcademicYear(
        id=str(uuid.uuid4()),
        tenant_id=tenant.id,
        name=name,
        start_date=date.fromisoformat(start),
        end_date=date.fromisoformat(end),
    )
    db_session.add(year)
    db_session.flush()
    return year


@pytest.fixture
def year(db_session, tenant):
    """One-month academic year (June 2026) for hand-checkable date math.

    June 2026 layout: Mon 1st … Tue 30th.
    Sundays: 7, 14, 21, 28. Saturdays: 6, 13, 20, 27 (2nd = 13, 4th = 27).
    """
    return _make_year(db_session, tenant)


def _ctx(flask_app, tenant):
    ctx = flask_app.test_request_context("/api/academics/calendar")
    ctx.push()
    g.tenant_id = tenant.id
    return ctx


@pytest.fixture
def request_ctx(flask_app, tenant):
    ctx = _ctx(flask_app, tenant)
    yield
    ctx.pop()


# ---------------------------------------------------------------------------
# Wizard state
# ---------------------------------------------------------------------------

def test_get_or_create_calendar_creates_then_reuses_draft(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    cal = services.get_or_create_calendar(year.id)
    assert cal.status == "draft"
    assert cal.current_step == 1
    assert cal.weekly_config == {
        "days": [], "second_saturday": False, "fourth_saturday": False,
    }

    again = services.get_or_create_calendar(year.id)
    assert again.id == cal.id


def test_get_or_create_calendar_rejects_unknown_year(db_session, tenant, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.services import CalendarValidationError

    with pytest.raises(CalendarValidationError):
        services.get_or_create_calendar(str(uuid.uuid4()))


def test_current_step_never_moves_backwards(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    cal = services.get_or_create_calendar(year.id)
    services.update_calendar(cal, {"current_step": 5})
    services.update_calendar(cal, {"current_step": 2})
    assert cal.current_step == 5


def test_weekly_config_syncs_recurring_holiday_rows(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.holidays import Holiday

    cal = services.get_or_create_calendar(year.id)
    services.update_calendar(
        cal,
        {"weekly_holidays_config": {"days": [6, 5], "second_saturday": True}},
    )

    rows = Holiday.query.filter(
        Holiday.tenant_id == tenant.id,
        Holiday.academic_year_id == year.id,
        Holiday.is_recurring.is_(True),
        Holiday.holiday_type == "weekly_off",
    ).all()
    assert sorted(r.recurring_day_of_week for r in rows) == [5, 6]

    # Deselecting Saturday removes its recurring row, keeps Sunday.
    services.update_calendar(cal, {"weekly_holidays_config": {"days": [6]}})
    rows = Holiday.query.filter(
        Holiday.tenant_id == tenant.id,
        Holiday.academic_year_id == year.id,
        Holiday.is_recurring.is_(True),
    ).all()
    assert [r.recurring_day_of_week for r in rows] == [6]


# ---------------------------------------------------------------------------
# Exam windows
# ---------------------------------------------------------------------------

def test_exam_window_must_fall_within_year(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.services import CalendarValidationError

    with pytest.raises(CalendarValidationError):
        services.create_exam_window(
            year.id,
            {"name": "Mid Term", "start_date": "2026-05-20", "end_date": "2026-06-05"},
        )


def test_exam_window_overlap_rejected_for_shared_scope(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.services import CalendarValidationError

    services.create_exam_window(
        year.id,
        {"name": "Mid Term", "start_date": "2026-06-08", "end_date": "2026-06-12",
         "applicable_class_ids": ["c1", "c2"]},
    )
    # Same dates, disjoint class scope → allowed.
    services.create_exam_window(
        year.id,
        {"name": "Junior Test", "start_date": "2026-06-10", "end_date": "2026-06-11",
         "applicable_class_ids": ["c9"]},
    )
    # Overlapping dates, shared class → rejected.
    with pytest.raises(CalendarValidationError):
        services.create_exam_window(
            year.id,
            {"name": "Retest", "start_date": "2026-06-11", "end_date": "2026-06-15",
             "applicable_class_ids": ["c2"]},
        )
    # Empty scope means all classes → also rejected.
    with pytest.raises(CalendarValidationError):
        services.create_exam_window(
            year.id,
            {"name": "All-school", "start_date": "2026-06-12", "end_date": "2026-06-12"},
        )


def test_exam_window_update_and_delete(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    window = services.create_exam_window(
        year.id,
        {"name": "Mid Term", "start_date": "2026-06-08", "end_date": "2026-06-12"},
    )
    updated = services.update_exam_window(window.id, {"name": "Mid Term Exam"})
    assert updated.name == "Mid Term Exam"

    assert services.delete_exam_window(window.id) is True
    assert services.list_exam_windows(year.id) == []


# ---------------------------------------------------------------------------
# School events
# ---------------------------------------------------------------------------

def test_school_event_validation_and_crud(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.services import CalendarValidationError

    with pytest.raises(CalendarValidationError):
        services.create_school_event(year.id, {"name": "", "event_date": "2026-06-10"})
    with pytest.raises(CalendarValidationError):
        services.create_school_event(
            year.id, {"name": "Sports Day", "event_date": "2026-07-10"}
        )

    event = services.create_school_event(
        year.id,
        {"name": "Sports Day", "event_type": "activity", "event_date": "2026-06-19"},
    )
    assert event.applies_to == "entire_school"

    updated = services.update_school_event(event.id, {"event_date": "2026-06-20"})
    assert updated.event_date == date(2026, 6, 20)

    assert services.delete_school_event(event.id) is True


# ---------------------------------------------------------------------------
# Working-day math + days feed
# ---------------------------------------------------------------------------

def _configured_calendar(services, year):
    """Sundays + 2nd/4th Saturday off, one public holiday, 3-day vacation."""
    cal = services.get_or_create_calendar(year.id)
    services.update_calendar(
        cal,
        {"weekly_holidays_config": {
            "days": [6], "second_saturday": True, "fourth_saturday": True,
        }},
    )
    return cal


def _add_holiday(db_session, tenant, year, name, start, end=None, holiday_type="public"):
    from modules.academics.calendar.holidays import Holiday

    h = Holiday(
        tenant_id=tenant.id,
        academic_year_id=year.id,
        name=name,
        holiday_type=holiday_type,
        start_date=date.fromisoformat(start),
        end_date=date.fromisoformat(end or start),
    )
    db_session.add(h)
    db_session.flush()
    return h


def test_summary_working_day_math(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    cal = _configured_calendar(services, year)
    _add_holiday(db_session, tenant, year, "Founders Day", "2026-06-10")
    _add_holiday(db_session, tenant, year, "Short Break", "2026-06-15", "2026-06-17",
                 holiday_type="vacation")

    summary = services.compute_summary(cal)

    # June 2026: 30 days. Weekly off: Sundays 7/14/21/28 + Sat 13 + Sat 27 = 6.
    # Public holiday June 10 = 1. Vacation June 15–17 = 3. Working = 30-10 = 20.
    assert summary["total_days"] == 30
    assert summary["weekly_holiday_days"] == 6
    assert summary["public_holiday_days"] == 1
    assert summary["vacation_days"] == 3
    assert summary["working_days"] == 20


def test_days_feed_classification_priority(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    cal = _configured_calendar(services, year)
    # Vacation covering a Sunday (June 14) must classify as vacation.
    _add_holiday(db_session, tenant, year, "Break", "2026-06-13", "2026-06-15",
                 holiday_type="vacation")
    _add_holiday(db_session, tenant, year, "Founders Day", "2026-06-10")

    feed = {d["date"]: d for d in services.get_days_feed(cal)}
    assert len(feed) == 30
    assert feed["2026-06-14"]["day_type"] == "vacation"       # vacation > weekly
    assert feed["2026-06-10"]["day_type"] == "public_holiday"
    assert feed["2026-06-07"]["day_type"] == "weekly_holiday"  # Sunday
    assert feed["2026-06-27"]["day_type"] == "weekly_holiday"  # 4th Saturday
    assert feed["2026-06-08"]["day_type"] == "working"


def test_days_feed_exam_and_event_markers(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    cal = _configured_calendar(services, year)
    services.create_exam_window(
        year.id,
        {"name": "Mid Term", "start_date": "2026-06-08", "end_date": "2026-06-09"},
    )
    services.create_school_event(
        year.id, {"name": "Sports Day", "event_date": "2026-06-19"}
    )

    feed = {d["date"]: d for d in services.get_days_feed(cal)}
    assert feed["2026-06-08"]["has_exam"] is True
    assert feed["2026-06-08"]["day_type"] == "working"  # exams don't stop work days
    assert feed["2026-06-19"]["has_event"] is True


# ---------------------------------------------------------------------------
# Publish
# ---------------------------------------------------------------------------

def _add_term(db_session, tenant, year, name, start, end, sequence=1):
    from modules.academics.backbone.models import AcademicTerm

    term = AcademicTerm(
        tenant_id=tenant.id,
        academic_year_id=year.id,
        name=name,
        sequence=sequence,
        start_date=date.fromisoformat(start),
        end_date=date.fromisoformat(end),
    )
    db_session.add(term)
    db_session.flush()
    return term


def test_publish_rejects_overlapping_terms(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.services import CalendarValidationError

    cal = _configured_calendar(services, year)
    _add_term(db_session, tenant, year, "Term 1", "2026-06-01", "2026-06-20")
    _add_term(db_session, tenant, year, "Term 2", "2026-06-15", "2026-06-30", sequence=2)

    with pytest.raises(CalendarValidationError):
        services.publish_calendar(cal, user_id=None)
    assert cal.status == "draft"


def test_publish_rejects_overlapping_vacations(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.services import CalendarValidationError

    cal = _configured_calendar(services, year)
    _add_holiday(db_session, tenant, year, "Break A", "2026-06-08", "2026-06-12",
                 holiday_type="vacation")
    _add_holiday(db_session, tenant, year, "Break B", "2026-06-11", "2026-06-16",
                 holiday_type="vacation")

    with pytest.raises(CalendarValidationError):
        services.publish_calendar(cal, user_id=None)


def test_publish_materializes_nth_saturdays_and_snapshots(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.holidays import Holiday

    cal = _configured_calendar(services, year)
    _add_term(db_session, tenant, year, "Term 1", "2026-06-01", "2026-06-15")
    _add_term(db_session, tenant, year, "Term 2", "2026-06-16", "2026-06-30", sequence=2)

    summary = services.publish_calendar(cal, user_id="user-1")

    assert cal.status == "published"
    assert cal.published_by == "user-1"
    assert cal.published_summary == summary
    assert summary["working_days"] == 24  # 30 - 4 Sundays - 2 Saturdays

    dated = Holiday.query.filter(
        Holiday.tenant_id == tenant.id,
        Holiday.academic_year_id == year.id,
        Holiday.is_recurring.is_(False),
        Holiday.holiday_type == "weekly_off",
    ).all()
    assert {(h.name, h.start_date.isoformat()) for h in dated} == {
        ("Second Saturday", "2026-06-13"),
        ("Fourth Saturday", "2026-06-27"),
    }

    # Re-publishing regenerates rather than duplicating the generated rows.
    services.publish_calendar(cal, user_id="user-1")
    dated_again = Holiday.query.filter(
        Holiday.tenant_id == tenant.id,
        Holiday.academic_year_id == year.id,
        Holiday.is_recurring.is_(False),
        Holiday.holiday_type == "weekly_off",
    ).count()
    assert dated_again == 2


# ---------------------------------------------------------------------------
# Dashboard additions: duplicates, semester markers, per-type stats, audit
# ---------------------------------------------------------------------------

def test_duplicate_event_name_on_same_date_rejected(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.services import CalendarValidationError

    services.create_school_event(
        year.id, {"name": "Sports Day", "event_date": "2026-06-19"}
    )
    with pytest.raises(CalendarValidationError):
        services.create_school_event(
            year.id, {"name": "sports day", "event_date": "2026-06-19"}
        )
    # Same name on a different date is fine.
    services.create_school_event(
        year.id, {"name": "Sports Day", "event_date": "2026-06-20"}
    )


def test_days_feed_semester_markers(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    cal = services.get_or_create_calendar(year.id)
    _add_term(db_session, tenant, year, "Term 1", "2026-06-01", "2026-06-15")

    feed = {d["date"]: d for d in services.get_days_feed(cal)}
    assert feed["2026-06-01"]["semester_start"] == "Term 1"
    assert feed["2026-06-15"]["semester_end"] == "Term 1"
    assert feed["2026-06-10"]["semester_start"] is None


def test_summary_events_by_type(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    cal = services.get_or_create_calendar(year.id)
    services.create_school_event(
        year.id, {"name": "Staff Training", "event_type": "training",
                  "event_date": "2026-06-19"}
    )
    services.create_school_event(
        year.id, {"name": "PTM", "event_type": "meeting", "event_date": "2026-06-20"}
    )
    summary = services.compute_summary(cal)
    assert summary["events_by_type"] == {"training": 1, "meeting": 1}


def test_event_mutations_write_audit_log(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.audit.models import TenantAuditLog

    event = services.create_school_event(
        year.id, {"name": "Annual Function", "event_date": "2026-06-19"}
    )
    services.update_school_event(event.id, {"event_date": "2026-06-20"})
    services.delete_school_event(event.id)

    actions = [
        row.action
        for row in TenantAuditLog.query.filter(
            TenantAuditLog.tenant_id == tenant.id,
            TenantAuditLog.resource_id == event.id,
        ).all()
    ]
    assert actions == [
        "school_event_created", "school_event_updated", "school_event_deleted",
    ]


def test_holiday_create_records_creator(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import holiday_services

    result = holiday_services.create_holiday(
        {"name": "Founders Day", "holiday_type": "public",
         "start_date": "2026-06-10", "academic_year_id": year.id},
        tenant.id,
    )
    assert result["success"] is True
    # No authenticated user in this context — created_by stays null but the
    # field is present in the payload for the dashboard.
    assert "created_by" in result["data"]


# ---------------------------------------------------------------------------
# Event status lifecycle (Story 3)
# ---------------------------------------------------------------------------

def test_event_status_defaults_active_and_validates(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.services import CalendarValidationError

    event = services.create_school_event(
        year.id, {"name": "Sports Day", "event_date": "2026-06-19"}
    )
    assert event.status == "active"

    with pytest.raises(CalendarValidationError):
        services.create_school_event(
            year.id,
            {"name": "Bad", "event_date": "2026-06-20", "status": "nonsense"},
        )


def test_non_active_events_excluded_from_summary_and_feed(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    cal = services.get_or_create_calendar(year.id)
    services.create_school_event(
        year.id, {"name": "Confirmed Day", "event_date": "2026-06-10"}
    )
    services.create_school_event(
        year.id,
        {"name": "Tentative Day", "event_date": "2026-06-11", "status": "draft"},
    )
    services.create_exam_window(
        year.id,
        {"name": "Cancelled Exam", "start_date": "2026-06-15",
         "end_date": "2026-06-16", "status": "cancelled"},
    )

    summary = services.compute_summary(cal)
    assert summary["event_count"] == 1        # draft excluded
    assert summary["exam_days"] == 0          # cancelled exam reserves no days

    feed = {d["date"]: d for d in services.get_days_feed(cal)}
    assert feed["2026-06-10"]["has_event"] is True
    assert feed["2026-06-11"]["has_event"] is False   # draft
    assert feed["2026-06-15"]["has_exam"] is False    # cancelled


def test_cancelled_exam_frees_overlapping_dates(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    first = services.create_exam_window(
        year.id, {"name": "Mid Term", "start_date": "2026-06-08", "end_date": "2026-06-12"}
    )
    # Overlapping window is normally rejected …
    services.update_exam_window(first.id, {"status": "cancelled"})
    # … but once the first is cancelled, its dates are free again.
    services.create_exam_window(
        year.id, {"name": "Mid Term Retry", "start_date": "2026-06-10", "end_date": "2026-06-14"}
    )


def test_cancelled_event_does_not_block_duplicate_name(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    event = services.create_school_event(
        year.id, {"name": "Sports Day", "event_date": "2026-06-19"}
    )
    services.update_school_event(event.id, {"status": "cancelled"})
    # Same name+date now allowed because the original is cancelled.
    services.create_school_event(
        year.id, {"name": "Sports Day", "event_date": "2026-06-19"}
    )


def test_event_name_and_description_length_validated(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.services import CalendarValidationError

    with pytest.raises(CalendarValidationError):
        services.create_school_event(
            year.id, {"name": "x" * 121, "event_date": "2026-06-19"}
        )
    with pytest.raises(CalendarValidationError):
        services.create_school_event(
            year.id,
            {"name": "Ok", "event_date": "2026-06-19", "description": "y" * 1001},
        )


# ---------------------------------------------------------------------------
# Tenant isolation
# ---------------------------------------------------------------------------

def test_calendar_not_visible_across_tenants(db_session, tenant, year, flask_app):
    from core.models import Tenant, TENANT_STATUS_ACTIVE, BILLING_CYCLE_YEARLY
    from modules.academics.calendar import services

    ctx = _ctx(flask_app, tenant)
    cal = services.get_or_create_calendar(year.id)
    cal_id = cal.id
    ctx.pop()

    other = Tenant(
        id=str(uuid.uuid4()),
        name="Other School",
        subdomain=f"other-{uuid.uuid4().hex}",
        status=TENANT_STATUS_ACTIVE,
        billing_cycle=BILLING_CYCLE_YEARLY,
    )
    db_session.add(other)
    db_session.flush()

    ctx = _ctx(flask_app, other)
    assert services.get_calendar(cal_id) is None
    assert services.get_calendar_for_year(year.id) is None
    ctx.pop()


# ---------------------------------------------------------------------------
# Term route validation helper (semesters)
# ---------------------------------------------------------------------------

def test_term_dates_validated_against_year_and_siblings(db_session, tenant, year, request_ctx):
    from modules.academics.term_routes import _validate_term_dates

    # Outside the year.
    err = _validate_term_dates(
        tenant.id, year.id, date(2026, 5, 20), date(2026, 6, 10)
    )
    assert err and "within the academic year" in err

    _add_term(db_session, tenant, year, "Term 1", "2026-06-01", "2026-06-15")

    # Overlapping sibling.
    err = _validate_term_dates(
        tenant.id, year.id, date(2026, 6, 10), date(2026, 6, 20)
    )
    assert err and "overlap" in err.lower()

    # Clean slot.
    assert _validate_term_dates(
        tenant.id, year.id, date(2026, 6, 16), date(2026, 6, 30)
    ) is None


# ---------------------------------------------------------------------------
# Vacation overlap via holidays service
# ---------------------------------------------------------------------------

def test_holiday_service_rejects_overlapping_vacations(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import holiday_services

    first = holiday_services.create_holiday(
        {"name": "Summer Break", "holiday_type": "vacation",
         "start_date": "2026-06-08", "end_date": "2026-06-12",
         "academic_year_id": year.id},
        tenant.id,
    )
    assert first["success"] is True

    second = holiday_services.create_holiday(
        {"name": "Extra Break", "holiday_type": "vacation",
         "start_date": "2026-06-11", "end_date": "2026-06-14",
         "academic_year_id": year.id},
        tenant.id,
    )
    assert second["success"] is False
    assert "overlap" in second["error"].lower()

    # Non-vacation holidays may still share those dates.
    normal = holiday_services.create_holiday(
        {"name": "Founders Day", "holiday_type": "public",
         "start_date": "2026-06-11", "academic_year_id": year.id},
        tenant.id,
    )
    assert normal["success"] is True


# ---------------------------------------------------------------------------
# Export (CSV / Excel / PDF)
# ---------------------------------------------------------------------------

def _full_calendar(db_session, tenant, year, services, holiday_services):
    """A calendar with one of every section, for export assertions."""
    cal = _configured_calendar(services, year)
    holiday_services.create_holiday(
        {"name": "Founders Day", "holiday_type": "public",
         "start_date": "2026-06-10", "academic_year_id": year.id},
        tenant.id,
    )
    holiday_services.create_holiday(
        {"name": "Summer Break", "holiday_type": "vacation",
         "start_date": "2026-06-15", "end_date": "2026-06-18",
         "academic_year_id": year.id},
        tenant.id,
    )
    _add_term(db_session, tenant, year, "Semester 1", "2026-06-01", "2026-06-30")
    services.create_exam_window(
        year.id,
        {"name": "Unit Test 1", "exam_type": "unit_test",
         "start_date": "2026-06-22", "end_date": "2026-06-24"},
    )
    services.create_school_event(
        year.id, {"name": "Sports Day", "event_type": "activity", "event_date": "2026-06-19"}
    )
    return cal


def test_build_export_payload_covers_all_sections(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services, export_services
    from modules.academics.calendar import holiday_services

    cal = _full_calendar(db_session, tenant, year, services, holiday_services)
    payload = export_services.build_export_payload(cal)

    assert payload["year_label"] == year.name
    assert payload["summary"]["semester_count"] == 1
    assert [h["name"] for h in payload["public_holidays"]] == ["Founders Day"]
    assert [v["name"] for v in payload["vacations"]] == ["Summer Break"]
    assert [s["name"] for s in payload["semesters"]] == ["Semester 1"]
    assert [x["name"] for x in payload["exam_windows"]] == ["Unit Test 1"]
    assert [e["name"] for e in payload["events"]] == ["Sports Day"]
    # Sunday is weekly-off (day 6) plus 2nd/4th Saturday.
    assert "Sunday" in payload["weekly_holidays"]


def test_export_csv_contains_every_section(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services, export_services
    from modules.academics.calendar import holiday_services

    cal = _full_calendar(db_session, tenant, year, services, holiday_services)
    content, mimetype, filename = export_services.export_calendar(cal, "csv")

    assert mimetype == "text/csv"
    assert filename.endswith(".csv")
    text = content.decode("utf-8-sig")
    for token in ("Overview", "Public Holidays", "Founders Day", "Vacations",
                  "Summer Break", "Semesters", "Examination Windows",
                  "Unit Test 1", "School Events", "Sports Day"):
        assert token in text


def test_export_excel_is_a_valid_workbook(db_session, tenant, year, request_ctx):
    from io import BytesIO
    from openpyxl import load_workbook
    from modules.academics.calendar import services, export_services
    from modules.academics.calendar import holiday_services

    cal = _full_calendar(db_session, tenant, year, services, holiday_services)
    content, mimetype, filename = export_services.export_calendar(cal, "excel")

    assert filename.endswith(".xlsx")
    assert content[:2] == b"PK"  # xlsx is a zip
    wb = load_workbook(BytesIO(content))
    assert "Overview" in wb.sheetnames
    assert "Public Holidays" in wb.sheetnames
    assert "School Events" in wb.sheetnames


def test_export_sections_filter_narrows_output(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services, export_services
    from modules.academics.calendar import holiday_services

    cal = _full_calendar(db_session, tenant, year, services, holiday_services)
    content, _, _ = export_services.export_calendar(cal, "csv", sections=["events"])
    text = content.decode("utf-8-sig")
    assert "Sports Day" in text          # requested section present
    assert "Founders Day" not in text    # public holidays excluded
    assert "Unit Test 1" not in text     # exam windows excluded


def test_export_rejects_unknown_format(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services, export_services
    from modules.academics.calendar import holiday_services

    cal = _full_calendar(db_session, tenant, year, services, holiday_services)
    with pytest.raises(ValueError):
        export_services.export_calendar(cal, "json")


def test_export_pdf_when_weasyprint_available(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services, export_services
    from modules.academics.calendar import holiday_services

    if not export_services.HAS_WEASYPRINT:
        pytest.skip("WeasyPrint not available in this environment")

    cal = _full_calendar(db_session, tenant, year, services, holiday_services)
    content, mimetype, filename = export_services.export_calendar(cal, "pdf")
    assert mimetype == "application/pdf"
    assert filename.endswith(".pdf")
    assert content[:4] == b"%PDF"


# ---------------------------------------------------------------------------
# Admin lifecycle: delete / archive / restore / preferences
# ---------------------------------------------------------------------------

def _published_calendar(db_session, services, year):
    cal = _configured_calendar(services, year)
    cal.status = "published"
    db_session.commit()
    return cal


def test_delete_draft_calendar(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    cal = services.get_or_create_calendar(year.id)
    services.delete_calendar(cal)
    assert services.get_calendar(cal.id) is None


def test_delete_rejects_published_calendar(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.services import CalendarValidationError

    cal = _published_calendar(db_session, services, year)
    with pytest.raises(CalendarValidationError):
        services.delete_calendar(cal)
    assert services.get_calendar(cal.id) is not None


def test_archive_then_restore(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    cal = _published_calendar(db_session, services, year)
    services.archive_calendar(cal, "user-1")
    assert cal.status == "archived"
    assert cal.archived_at is not None
    assert cal.archived_by == "user-1"

    services.restore_calendar(cal, "user-1")
    assert cal.status == "published"
    assert cal.archived_at is None


def test_archive_rejects_draft(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.services import CalendarValidationError

    cal = services.get_or_create_calendar(year.id)
    with pytest.raises(CalendarValidationError):
        services.archive_calendar(cal, "user-1")


def test_update_preferences_valid_and_invalid(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.academics.calendar.services import CalendarValidationError

    cal = services.get_or_create_calendar(year.id)
    assert cal.prefs["default_view"] == "month"  # default

    services.update_preferences(cal, {"default_view": "week", "week_start": "sunday",
                                       "default_month": "2026-06"})
    assert cal.preferences["default_view"] == "week"
    assert cal.preferences["week_start"] == "sunday"
    assert cal.preferences["default_month"] == "2026-06"

    with pytest.raises(CalendarValidationError):
        services.update_preferences(cal, {"default_view": "spiral"})
    with pytest.raises(CalendarValidationError):
        services.update_preferences(cal, {"default_month": "June"})


# ---------------------------------------------------------------------------
# Import (template + row validation)
# ---------------------------------------------------------------------------

def test_import_template_has_header_and_sample(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import import_services

    csv_bytes = import_services.import_template_csv("events")
    text = csv_bytes.decode("utf-8-sig")
    assert "name,event_type,event_date" in text.splitlines()[0]
    assert len(text.strip().splitlines()) == 2  # header + one sample


def test_import_events_valid_and_invalid_rows(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services, import_services

    cal = services.get_or_create_calendar(year.id)
    csv_text = (
        "name,event_type,event_date,applies_to,description\n"
        "Sports Day,activity,2026-06-19,entire_school,Fun\n"
        "Bad Row,event,not-a-date,,\n"
        "Annual Day,celebration,2026-06-25,,\n"
    )
    report = import_services.import_rows(cal, "events", csv_text)
    assert report["total"] == 3
    assert report["imported"] == 2
    assert report["skipped"] == 1
    assert report["errors"][0]["row"] == 3
    assert report["errors"][0]["field"] == "event_date"

    names = {e.name for e in services.list_school_events(year.id)}
    assert {"Sports Day", "Annual Day"} <= names
    assert "Bad Row" not in names


def test_import_holidays_reports_duplicate(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services, import_services

    cal = _configured_calendar(services, year)
    csv_text = (
        "name,holiday_type,start_date,end_date,applies_to,description\n"
        "Republic Day,national,2026-06-10,,entire_school,\n"
        "Republic Day,national,2026-06-10,,entire_school,\n"  # duplicate name+date
    )
    report = import_services.import_rows(cal, "public_holidays", csv_text)
    assert report["imported"] == 1
    assert report["skipped"] == 1
    assert report["errors"][0]["row"] == 3


def test_import_rejects_missing_required_column(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services, import_services

    cal = services.get_or_create_calendar(year.id)
    with pytest.raises(import_services.ImportValidationError):
        import_services.import_rows(cal, "events", "name,notes\nSports Day,fun\n")


def test_import_rejects_unknown_type(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services, import_services

    cal = services.get_or_create_calendar(year.id)
    with pytest.raises(import_services.ImportValidationError):
        import_services.import_rows(cal, "teachers", "name\nx\n")


# ---------------------------------------------------------------------------
# Audit enrichment + activity feed
# ---------------------------------------------------------------------------

def test_event_update_records_changed_fields_diff(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services
    from modules.audit.models import TenantAuditLog

    ev = services.create_school_event(
        year.id, {"name": "Sports Day", "event_date": "2026-06-19"}
    )
    services.update_school_event(
        ev.id, {"name": "Annual Sports Day", "event_date": "2026-06-20"}
    )

    log = (
        TenantAuditLog.query.filter_by(
            tenant_id=tenant.id, action="school_event_updated"
        )
        .order_by(TenantAuditLog.created_at.desc())
        .first()
    )
    assert log is not None
    changes = (log.meta or {}).get("changes", {})
    assert changes["name"]["from"] == "Sports Day"
    assert changes["name"]["to"] == "Annual Sports Day"
    assert changes["event_date"]["from"] == "2026-06-19"


def test_activity_feed_newest_first(db_session, tenant, year, request_ctx):
    from modules.academics.calendar import services

    cal = services.get_or_create_calendar(year.id)
    services.create_school_event(year.id, {"name": "Event A", "event_date": "2026-06-10"})
    services.record_calendar_activity(
        cal, "export_completed", "Calendar exported as PDF", {"format": "pdf"}
    )

    feed = services.list_calendar_activity(cal, page=1, page_size=50)
    actions = [i["action"] for i in feed["items"]]
    assert "export_completed" in actions
    assert "school_event_created" in actions
    # Latest activity first.
    assert feed["items"][0]["action"] == "export_completed"
    assert feed["pagination"]["total_items"] >= 2
