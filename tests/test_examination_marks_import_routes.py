"""Uploading a marks sheet over REST.

The service was verified in EX-02B, so what is tested here is the transport
around it: that authority, the paper's lock and tenancy are enforced through
the HTTP door too, that a preview writes nothing, and that a refused import
returns the row report a teacher needs rather than a bare failure.
"""

from __future__ import annotations

import uuid
from datetime import date
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from modules.examinations import marks_service
from modules.examinations.models import MARK_ABSENT, MARK_PRESENT, ExamMark

from tests.test_examination_marks import (  # noqa: F401
    _new_id,
    ctx,
    cycles,
    exam_type,
    school,
    year,
)
from tests.test_examination_graphql import (  # noqa: F401
    EXAMINATIONS_ON,
    examinations_enabled,
    _read_as,
    _signed_in,
    _signed_in_at,
    client,
)
from tests.test_examination_results import _exam, _paper_spec, _papers  # noqa: F401

PREVIEW = "/api/examinations/papers/{}/marks/preview"
IMPORT = "/api/examinations/papers/{}/marks/import"
TEMPLATE = "/api/examinations/papers/{}/marks/template"


def _sheet(rows, *, headers=("admission_number", "marks", "status")):
    book = Workbook()
    sheet = book.active
    sheet.append(list(headers))
    for row in rows:
        sheet.append(list(row))
    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _headers(tenant, token):
    return {
        "X-Tenant-Subdomain": tenant.subdomain,
        "Authorization": f"Bearer {token}",
    }


def _upload(client, tenant, token, url, content, filename="marks.xlsx"):
    return client.post(
        url,
        data={"file": (BytesIO(content), filename)},
        content_type="multipart/form-data",
        headers=_headers(tenant, token),
    )


@pytest.fixture
def marker(db_session, tenant):
    return _signed_in(
        db_session, tenant,
        permissions=["assessment.enter", "assessment.manage", "assessment.read.class"],
    )


@pytest.fixture
def open_paper(ctx, tenant, cycles, exam_type, school):
    examination = _exam(
        tenant, cycles, exam_type, school,
        [_paper_spec(school["maths"], 100, pass_marks=35)],
    )
    return _papers(examination, tenant)[0]


def _marks(tenant, paper):
    _read_as(tenant)
    return ExamMark.query.filter_by(exam_paper_id=paper.id).all()


# ---------------------------------------------------------------------------
# Preview writes nothing
# ---------------------------------------------------------------------------

def test_a_valid_sheet_previews_without_writing(
    client, tenant, school, open_paper, marker, ctx
):
    _user, token = marker
    response = _upload(client, tenant, token, PREVIEW.format(open_paper.id), _sheet([
        (school["riya"].admission_number, 88, "present"),
        (school["dev"].admission_number, None, "absent"),
    ]))

    assert response.status_code == 200, response.get_json()
    body = response.get_json()["data"]
    assert body["summary"] == {"valid": 2, "invalid": 0, "total": 2}
    assert [row["valid"] for row in body["preview"]] == [True, True]
    assert _marks(tenant, open_paper) == []


def test_an_invalid_preview_reports_every_row(
    client, tenant, school, open_paper, marker, ctx
):
    """A teacher fixing a sheet needs the whole list, not the first problem."""
    _user, token = marker
    response = _upload(client, tenant, token, PREVIEW.format(open_paper.id), _sheet([
        (school["riya"].admission_number, 88, "present"),
        ("GHOST", 40, "present"),
        (school["dev"].admission_number, 5000, "present"),
    ]))

    body = response.get_json()["data"]
    assert body["summary"] == {"valid": 1, "invalid": 2, "total": 3}
    assert any("STUDENT_NOT_ELIGIBLE" in e for e in body["preview"][1]["errors"])
    assert any("MARKS_ABOVE_MAX" in e for e in body["preview"][2]["errors"])
    assert [row["row_number"] for row in body["preview"]] == [2, 3, 4]
    assert _marks(tenant, open_paper) == []


# ---------------------------------------------------------------------------
# Import is all or nothing
# ---------------------------------------------------------------------------

def test_a_valid_sheet_imports_every_row(
    client, tenant, school, open_paper, marker, ctx
):
    _user, token = marker
    response = _upload(client, tenant, token, IMPORT.format(open_paper.id), _sheet([
        (school["riya"].admission_number, 88, "present"),
        (school["dev"].admission_number, 0, "present"),
        (school["meera"].admission_number, None, "absent"),
    ]))

    assert response.status_code == 200, response.get_json()
    assert response.get_json()["data"]["imported"] == 3

    written = {m.student_id: m for m in _marks(tenant, open_paper)}
    assert len(written) == 3
    assert float(written[school["riya"].id].marks_obtained) == 88.0
    # A genuine zero survives the sheet as present-with-zero.
    assert written[school["dev"].id].status == MARK_PRESENT
    assert float(written[school["dev"].id].marks_obtained) == 0.0
    assert written[school["meera"].id].status == MARK_ABSENT
    assert written[school["meera"].id].marks_obtained is None


def test_valid_valid_invalid_valid_writes_nothing(
    client, tenant, school, open_paper, marker, ctx
):
    """The mandated atomicity case."""
    _user, token = marker
    response = _upload(client, tenant, token, IMPORT.format(open_paper.id), _sheet([
        (school["riya"].admission_number, 88, "present"),
        (school["dev"].admission_number, 70, "present"),
        ("GHOST", 40, "present"),
        (school["meera"].admission_number, 71, "present"),
    ]))

    assert response.status_code == 400
    body = response.get_json()
    assert body["details"]["code"] == "ROWS_INVALID"
    # The report travels with the refusal.
    assert body["details"]["summary"] == {"valid": 3, "invalid": 1, "total": 4}
    assert len(body["details"]["preview"]) == 4
    assert _marks(tenant, open_paper) == []


def test_an_existing_mark_stops_the_whole_import(
    client, tenant, school, open_paper, marker, ctx
):
    """Import creates; it never overwrites."""
    _user, token = marker
    _read_as(tenant)
    marks_service.record_mark(
        tenant_id=tenant.id, exam_paper_id=open_paper.id,
        student_id=school["riya"].id, status=MARK_PRESENT, marks_obtained=61,
        actor_user_id=_user.id, commit=False,
    )

    response = _upload(client, tenant, token, IMPORT.format(open_paper.id), _sheet([
        (school["riya"].admission_number, 88, "present"),
        (school["dev"].admission_number, 70, "present"),
    ]))

    assert response.status_code == 400
    assert any(
        "ALREADY_MARKED" in e
        for e in response.get_json()["details"]["preview"][0]["errors"]
    )
    remaining = _marks(tenant, open_paper)
    assert len(remaining) == 1
    assert float(remaining[0].marks_obtained) == 61.0


def test_a_repeated_admission_number_is_refused(
    client, tenant, school, open_paper, marker, ctx
):
    _user, token = marker
    number = school["riya"].admission_number
    response = _upload(client, tenant, token, IMPORT.format(open_paper.id), _sheet([
        (number, 88, "present"),
        (number, 44, "present"),
    ]))
    assert any(
        "STUDENT_REPEATED" in e
        for e in response.get_json()["details"]["preview"][1]["errors"]
    )
    assert _marks(tenant, open_paper) == []


@pytest.mark.parametrize(
    "marks,status,code",
    [
        ("nan", "present", "MARKS_INVALID"),
        ("inf", "present", "MARKS_INVALID"),
        ("1e999", "present", "MARKS_INVALID"),
        (-1, "present", "MARKS_INVALID"),
        (None, "present", "MARKS_REQUIRED"),
        (40, "absent", "MARKS_NOT_ALLOWED"),
        (40, "maybe", "STATUS_INVALID"),
    ],
    ids=["nan", "inf", "overflow", "negative", "no-mark", "absent-with-mark", "bad-status"],
)
def test_a_row_obeys_the_manual_entry_rules(
    client, tenant, school, open_paper, marker, ctx, marks, status, code
):
    _user, token = marker
    response = _upload(client, tenant, token, IMPORT.format(open_paper.id), _sheet([
        (school["riya"].admission_number, marks, status),
    ]))
    assert response.status_code == 400
    assert any(
        code in e for e in response.get_json()["details"]["preview"][0]["errors"]
    )
    assert _marks(tenant, open_paper) == []


def test_a_batch_students_additional_enrollment_imports(
    ctx, client, db_session, tenant, cycles, exam_type, school, marker
):
    """EX-02A.1's cohort rule reaches the importer: an `additional` enrollment
    is eligibility like any other."""
    from modules.examinations import services as exam_services

    created = exam_services.create_examination(
        tenant_id=tenant.id, academic_cycle_id=cycles["batch"].id,
        exam_type_id=exam_type.id, name=f"JEE-{uuid.uuid4().hex[:5]}",
        papers=[{"class_subject_id": school["jee_physics"].id,
                 "max_marks": 300, "exam_date": date(2026, 5, 20)}],
        commit=False,
    )
    examination = created["examination"]
    exam_services.schedule_examination(examination.id, tenant.id, commit=False)
    exam_services.open_marks_entry(examination.id, tenant.id, commit=False)
    paper = _papers(examination, tenant)[0]

    _user, token = marker
    response = _upload(client, tenant, token, IMPORT.format(paper.id), _sheet([
        (school["meera"].admission_number, 250, "present"),
    ]))
    assert response.status_code == 200, response.get_json()
    assert len(_marks(tenant, paper)) == 1


# ---------------------------------------------------------------------------
# The file itself
# ---------------------------------------------------------------------------

def test_a_missing_or_wrong_file_is_refused(
    client, tenant, open_paper, marker, ctx
):
    _user, token = marker
    url = IMPORT.format(open_paper.id)

    assert client.post(
        url, data={}, content_type="multipart/form-data",
        headers=_headers(tenant, token),
    ).status_code == 400

    assert _upload(
        client, tenant, token, url, b"a,b\n1,2\n", filename="marks.csv"
    ).status_code == 400

    assert _upload(client, tenant, token, url, b"").status_code == 400


def test_a_header_only_sheet_reports_no_rows_rather_than_success(
    client, tenant, open_paper, marker, ctx
):
    _user, token = marker
    response = _upload(client, tenant, token, IMPORT.format(open_paper.id), _sheet([]))
    assert response.status_code == 400
    assert response.get_json()["details"]["code"] == "NO_ROWS"


def test_a_sheet_without_admission_numbers_is_refused(
    client, tenant, open_paper, marker, ctx
):
    _user, token = marker
    response = _upload(
        client, tenant, token, IMPORT.format(open_paper.id),
        _sheet([(88,)], headers=("marks",)),
    )
    assert response.get_json()["details"]["code"] == "COLUMNS_MISSING"


# ---------------------------------------------------------------------------
# Authority, lock and tenancy
# ---------------------------------------------------------------------------

def test_uploading_needs_the_enter_key(
    client, tenant, db_session, school, open_paper, ctx
):
    _user, token = _signed_in(db_session, tenant, permissions=["examination.read"])
    response = _upload(client, tenant, token, IMPORT.format(open_paper.id), _sheet([
        (school["riya"].admission_number, 88, "present"),
    ]))
    assert response.status_code == 403
    assert _marks(tenant, open_paper) == []


def test_holding_the_key_without_standing_over_the_paper_is_refused(
    client, tenant, db_session, school, open_paper, ctx
):
    """ADR-014 still decides, through the service — the route does not restate it."""
    _user, token = _signed_in(db_session, tenant, permissions=["assessment.enter"])
    response = _upload(client, tenant, token, IMPORT.format(open_paper.id), _sheet([
        (school["riya"].admission_number, 88, "present"),
    ]))
    assert response.status_code == 403
    assert response.get_json()["details"]["code"] == "NOT_THE_MARKER"
    assert _marks(tenant, open_paper) == []


def test_a_locked_paper_refuses_an_upload(
    client, tenant, school, open_paper, marker, ctx
):
    _user, token = marker
    _read_as(tenant)
    marks_service.lock_paper(
        open_paper.id, tenant.id, actor_user_id=_user.id, commit=False
    )

    response = _upload(client, tenant, token, IMPORT.format(open_paper.id), _sheet([
        (school["riya"].admission_number, 88, "present"),
    ]))
    assert response.status_code == 409
    assert response.get_json()["details"]["code"] == "PAPER_LOCKED"
    assert _marks(tenant, open_paper) == []


def test_a_paper_this_school_does_not_have_is_not_found(
    client, tenant, school, open_paper, marker, ctx
):
    """The paper lookup is tenant-scoped, so another school's id is simply not
    there — and the refusal says so without confirming it exists elsewhere."""
    _user, token = marker
    response = _upload(
        client, tenant, token, IMPORT.format("p-someone-elses"),
        _sheet([(school["riya"].admission_number, 88, "present")]),
    )
    assert response.status_code == 404
    assert response.get_json()["details"]["code"] == "PAPER_NOT_FOUND"
    assert _marks(tenant, open_paper) == []


def test_an_account_of_another_school_is_refused_before_the_paper(
    client, db_session, tenant, school, open_paper, ctx
):
    """A foreign signed-in account never reaches the import at all.

    It is refused at authentication rather than at the paper lookup — the
    account is not one of this school's — so the assertion is that it is
    refused and writes nothing, not which of the two guards caught it.
    """
    from core.models import BILLING_CYCLE_YEARLY, TENANT_STATUS_ACTIVE, Tenant

    other = Tenant(
        id=_new_id("t-"), name="Other", subdomain=f"o-{uuid.uuid4().hex[:10]}",
        status=TENANT_STATUS_ACTIVE, billing_cycle=BILLING_CYCLE_YEARLY,
        feature_flags=EXAMINATIONS_ON,
    )
    db_session.add(other)
    db_session.flush()
    _outsider, outsider_token = _signed_in_at(
        db_session, other, permissions=["assessment.enter", "assessment.manage"]
    )

    response = _upload(
        client, other, outsider_token, IMPORT.format(open_paper.id),
        _sheet([(school["riya"].admission_number, 88, "present")]),
    )
    assert response.status_code >= 400
    assert _marks(tenant, open_paper) == []


# ---------------------------------------------------------------------------
# Template
# ---------------------------------------------------------------------------

def test_the_template_carries_the_register_and_no_marks(
    client, tenant, school, open_paper, marker, ctx
):
    _user, token = marker
    response = client.get(
        TEMPLATE.format(open_paper.id), headers=_headers(tenant, token)
    )
    assert response.status_code == 200
    assert "attachment" in response.headers["Content-Disposition"]

    book = load_workbook(BytesIO(response.data))
    rows = list(book.active.iter_rows(values_only=True))
    assert rows[0] == ("admission_number", "marks", "status")
    numbers = {row[0] for row in rows[1:]}
    assert numbers == {
        school["riya"].admission_number,
        school["dev"].admission_number,
        school["meera"].admission_number,
    }
    # No marks are pre-filled: the sheet is for entering, not for re-uploading.
    assert all(row[1] is None and row[2] is None for row in rows[1:])


def test_the_template_answers_to_the_same_authority(
    client, tenant, db_session, open_paper, ctx
):
    _user, token = _signed_in(db_session, tenant, permissions=["examination.read"])
    response = client.get(
        TEMPLATE.format(open_paper.id), headers=_headers(tenant, token)
    )
    assert response.status_code == 403


def test_the_round_trip_a_teacher_actually_performs(
    client, tenant, school, open_paper, marker, ctx
):
    """Download the template, fill it in, preview, import, read it back."""
    _user, token = marker
    downloaded = client.get(
        TEMPLATE.format(open_paper.id), headers=_headers(tenant, token)
    )
    book = load_workbook(BytesIO(downloaded.data))
    sheet = book.active
    for index, marks in enumerate([78, 0, None], start=2):
        sheet.cell(row=index, column=2, value=marks)
        sheet.cell(row=index, column=3, value="present" if marks is not None else "absent")
    buffer = BytesIO()
    book.save(buffer)
    filled = buffer.getvalue()

    previewed = _upload(client, tenant, token, PREVIEW.format(open_paper.id), filled)
    assert previewed.get_json()["data"]["summary"]["invalid"] == 0
    assert _marks(tenant, open_paper) == []

    imported = _upload(client, tenant, token, IMPORT.format(open_paper.id), filled)
    assert imported.status_code == 200, imported.get_json()
    assert imported.get_json()["data"]["imported"] == 3

    _read_as(tenant)
    register = marks_service.marking_register(open_paper.id, tenant.id)
    states = {
        row["admission_number"]: (row["status"], row["marks_obtained"])
        for row in register["students"]
    }
    assert states[school["riya"].admission_number] == ("present", 78.0)
    assert states[school["dev"].admission_number] == ("present", 0.0)
    assert states[school["meera"].admission_number] == ("absent", None)
    assert register["progress"]["recorded"] == 3
