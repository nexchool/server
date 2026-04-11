#!/usr/bin/env python3
"""
Generate sample .xlsx files for bulk student import QA.

Usage (from repo root or server/):
  python server/scripts/generate_bulk_import_sample_xlsx.py

Requires: openpyxl
Outputs:
  docs/bulk-import-samples/bulk-import-european-10-students.xlsx
  docs/bulk-import-samples/bulk-import-indian-100-mixed.xlsx

Use academic year that has classes "Grade 10" + section "A" or "B" (matches common seed data).
"""

from __future__ import annotations

import random
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

# Align with bulk import normalized headers (snake_case)
HEADERS = [
    "name",
    "email",
    "admission_number",
    "class_name",
    "section",
    "roll_number",
    "gender",
    "date_of_birth",
    "phone",
    "father_name",
    "father_phone",
    "father_email",
    "father_occupation",
    "father_annual_income",
    "mother_name",
    "mother_phone",
    "mother_email",
    "mother_occupation",
    "mother_annual_income",
    "guardian_name",
    "guardian_phone",
    "guardian_email",
    "guardian_relationship",
    "current_address",
    "current_city",
    "current_state",
    "current_pincode",
    "permanent_address",
    "permanent_city",
    "permanent_state",
    "permanent_pincode",
    "is_same_as_permanent_address",
    "aadhar_number",
    "apaar_id",
    "emis_number",
    "udise_student_id",
    "religion",
    "category",
    "caste",
    "nationality",
    "mother_tongue",
    "place_of_birth",
    "blood_group",
    "height_cm",
    "weight_kg",
    "medical_allergies",
    "medical_conditions",
    "emergency_contact_name",
    "emergency_contact_phone",
    "emergency_contact_relationship",
    "admission_date",
    "previous_school_name",
    "previous_school_class",
    "last_school_board",
    "tc_number",
    "house_name",
    "student_status",
    "is_transport_opted",
]


def col_index(name: str) -> int:
    return HEADERS.index(name) + 1


def write_headers(ws) -> None:
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)


def build_row(**kwargs) -> list:
    row = [""] * len(HEADERS)
    for k, v in kwargs.items():
        if k in HEADERS:
            row[HEADERS.index(k)] = v
    return row


def save_wb(wb: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {path}")


def european_10() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    write_headers(ws)

    # Realistic EU-style names, addresses, phones; Grade 10 A/B matches common seed.
    rows = [
        build_row(
            name="Lars Eriksson",
            email="lars.eriksson.import01@studentmail.eu",
            admission_number="EU-ADM-2026-001",
            class_name="Grade 10",
            section="A",
            roll_number=1,
            gender="Male",
            date_of_birth="2008-03-14",
            phone="+46 70 123 45 67",
            father_name="Erik Eriksson",
            father_phone="+46 70 234 56 78",
            father_email="erik.eriksson@mail.se",
            father_occupation="Civil engineer",
            father_annual_income=72000,
            mother_name="Anna Eriksson",
            mother_phone="+46 70 345 67 89",
            mother_email="anna.eriksson@mail.se",
            mother_occupation="Nurse",
            mother_annual_income=58000,
            current_address="Vasagatan 12",
            current_city="Stockholm",
            current_state="Stockholm",
            current_pincode="111 20",
            permanent_address="Vasagatan 12",
            permanent_city="Stockholm",
            permanent_state="Stockholm",
            permanent_pincode="111 20",
            is_same_as_permanent_address=True,
            nationality="Swedish",
            mother_tongue="Swedish",
            place_of_birth="Stockholm",
            blood_group="A+",
            height_cm=172,
            weight_kg=61.5,
            medical_allergies="",
            medical_conditions="",
            emergency_contact_name="Erik Eriksson",
            emergency_contact_phone="+46 70 234 56 78",
            emergency_contact_relationship="Father",
            admission_date="2024-06-01",
            previous_school_name="Östra Real School",
            previous_school_class="Grade 9",
            last_school_board="Swedish National",
            student_status="active",
            is_transport_opted=False,
        ),
        build_row(
            name="Sofia Müller",
            email="sofia.mueller.import02@schule-mail.de",
            admission_number="EU-ADM-2026-002",
            class_name="Grade 10",
            section="A",
            roll_number=2,
            gender="Female",
            date_of_birth="2008-07-22",
            phone="+49 151 22334455",
            father_name="Thomas Müller",
            father_phone="+49 151 33445566",
            father_email="thomas.mueller@example.de",
            father_occupation="Software architect",
            father_annual_income=88000,
            mother_name="Claudia Müller",
            mother_phone="+49 151 44556677",
            mother_email="claudia.mueller@example.de",
            mother_occupation="Pharmacist",
            mother_annual_income=62000,
            current_address="Schönhauser Allee 45",
            current_city="Berlin",
            current_state="Berlin",
            current_pincode="10437",
            nationality="German",
            mother_tongue="German",
            place_of_birth="Munich",
            blood_group="O+",
            height_cm=165,
            weight_kg=54.0,
            emergency_contact_name="Thomas Müller",
            emergency_contact_phone="+49 151 33445566",
            emergency_contact_relationship="Father",
            admission_date="2024-06-01",
            previous_school_name="Gymnasium München",
            house_name="House Cerulean",
            student_status="active",
            is_transport_opted=True,
        ),
        build_row(
            name="Lucas van Dijk",
            email="lucas.vandijk.import03@school.nl",
            admission_number="EU-ADM-2026-003",
            class_name="Grade 10",
            section="B",
            roll_number=1,
            gender="Male",
            date_of_birth="2008-11-05",
            phone="+31 6 12345678",
            father_name="Willem van Dijk",
            father_phone="+31 6 87654321",
            father_email="w.vandijk@ziggo.nl",
            father_occupation="Logistics manager",
            father_annual_income=69000,
            mother_name="Emma van Dijk",
            mother_phone="+31 6 11223344",
            mother_email="emma.vd@ziggo.nl",
            mother_occupation="Teacher",
            mother_annual_income=52000,
            current_address="Keizersgracht 210",
            current_city="Amsterdam",
            current_state="North Holland",
            current_pincode="1016",
            religion="",
            nationality="Dutch",
            mother_tongue="Dutch",
            place_of_birth="Rotterdam",
            blood_group="B+",
            height_cm=178,
            weight_kg=70.0,
            medical_allergies="Peanuts",
            emergency_contact_name="Emma van Dijk",
            emergency_contact_phone="+31 6 11223344",
            emergency_contact_relationship="Mother",
            admission_date="2024-06-03",
            student_status="active",
            is_transport_opted=False,
        ),
        build_row(
            name="Chiara Rossi",
            email="chiara.rossi.import04@liceo.it",
            admission_number="EU-ADM-2026-004",
            class_name="Grade 10",
            section="B",
            roll_number=2,
            gender="Female",
            date_of_birth="2008-01-30",
            phone="+39 320 1122334",
            father_name="Marco Rossi",
            father_phone="+39 320 2233445",
            father_email="marco.rossi@tin.it",
            father_occupation="Architect",
            father_annual_income=55000,
            mother_name="Giulia Rossi",
            mother_phone="+39 320 3344556",
            mother_email="giulia.rossi@tin.it",
            mother_occupation="Dentist",
            mother_annual_income=61000,
            current_address="Via Garibaldi 8",
            current_city="Milan",
            current_state="Lombardy",
            current_pincode="20121",
            nationality="Italian",
            mother_tongue="Italian",
            place_of_birth="Milan",
            blood_group="AB+",
            height_cm=162,
            weight_kg=52.0,
            admission_date="2024-06-01",
            last_school_board="Italian Ministry curriculum",
            student_status="active",
            is_transport_opted=False,
        ),
        build_row(
            name="Oliver Hughes",
            email="oliver.hughes.import05@schools.wales.uk",
            admission_number="EU-ADM-2026-005",
            class_name="Grade 10",
            section="A",
            roll_number=3,
            gender="Male",
            date_of_birth="2008-09-12",
            phone="+44 7700 900123",
            father_name="David Hughes",
            father_phone="+44 7700 900456",
            father_email="david.hughes@mail.co.uk",
            father_occupation="NHS consultant",
            father_annual_income=95000,
            mother_name="Sarah Hughes",
            mother_phone="+44 7700 900789",
            mother_email="sarah.hughes@mail.co.uk",
            mother_occupation="Solicitor",
            mother_annual_income=78000,
            current_address="14 Cathedral Road",
            current_city="Cardiff",
            current_state="Wales",
            current_pincode="CF11 9LJ",
            nationality="British",
            mother_tongue="English",
            place_of_birth="Cardiff",
            blood_group="O-",
            height_cm=175,
            weight_kg=68.0,
            emergency_contact_name="Sarah Hughes",
            emergency_contact_phone="+44 7700 900789",
            emergency_contact_relationship="Mother",
            admission_date="2024-06-01",
            student_status="active",
            is_transport_opted=True,
        ),
        build_row(
            name="Amélie Dupont",
            email="amelie.dupont.import06@lycee.fr",
            admission_number="EU-ADM-2026-006",
            class_name="Grade 10",
            section="A",
            roll_number=4,
            gender="Female",
            date_of_birth="2008-04-18",
            phone="+33 6 12 34 56 78",
            father_name="Jean Dupont",
            father_phone="+33 6 23 45 67 89",
            father_email="jean.dupont@orange.fr",
            father_occupation="Bank officer",
            father_annual_income=58000,
            mother_name="Camille Dupont",
            mother_phone="+33 6 34 56 78 90",
            mother_email="camille.dupont@orange.fr",
            mother_occupation="Graphic designer",
            mother_annual_income=48000,
            current_address="Rue de Rivoli 90",
            current_city="Paris",
            current_state="Île-de-France",
            current_pincode="75004",
            nationality="French",
            mother_tongue="French",
            place_of_birth="Lyon",
            blood_group="A-",
            height_cm=160,
            weight_kg=50.0,
            admission_date="2024-06-01",
            student_status="active",
            is_transport_opted=False,
        ),
        build_row(
            name="Mikołaj Kowalski",
            email="mikolaj.kowalski.import07@szkola.pl",
            admission_number="EU-ADM-2026-007",
            class_name="Grade 10",
            section="B",
            roll_number=3,
            gender="Male",
            date_of_birth="2008-06-02",
            phone="+48 512 345 678",
            father_name="Piotr Kowalski",
            father_phone="+48 601 234 567",
            father_email="piotr.kowalski@wp.pl",
            father_occupation="Electrician",
            father_annual_income=42000,
            mother_name="Magda Kowalski",
            mother_phone="+48 602 345 678",
            mother_email="magda.kowalski@wp.pl",
            mother_occupation="Retail manager",
            mother_annual_income=38000,
            current_address="ul. Marszałkowska 100",
            current_city="Warsaw",
            current_state="Masovian",
            current_pincode="00-102",
            nationality="Polish",
            mother_tongue="Polish",
            place_of_birth="Warsaw",
            blood_group="B-",
            height_cm=180,
            weight_kg=72.0,
            admission_date="2024-06-01",
            student_status="active",
            is_transport_opted=False,
        ),
        build_row(
            name="Elena Vasquez",
            email="elena.vasquez.import08@colegio.es",
            admission_number="EU-ADM-2026-008",
            class_name="Grade 10",
            section="A",
            roll_number=5,
            gender="Female",
            date_of_birth="2008-12-20",
            phone="+34 612 345 678",
            father_name="Carlos Vasquez",
            father_phone="+34 623 456 789",
            father_email="c.vasquez@movistar.es",
            father_occupation="Hotel director",
            father_annual_income=51000,
            mother_name="Maria Vasquez",
            mother_phone="+34 634 567 890",
            mother_email="m.vasquez@movistar.es",
            mother_occupation="Physiotherapist",
            mother_annual_income=44000,
            current_address="Calle de Alcalá 400",
            current_city="Madrid",
            current_state="Madrid",
            current_pincode="28027",
            nationality="Spanish",
            mother_tongue="Spanish",
            place_of_birth="Madrid",
            blood_group="A+",
            height_cm=163,
            weight_kg=53.5,
            medical_conditions="Mild asthma (inhaler)",
            emergency_contact_name="Maria Vasquez",
            emergency_contact_phone="+34 634 567 890",
            emergency_contact_relationship="Mother",
            admission_date="2024-06-01",
            student_status="active",
            is_transport_opted=True,
        ),
        build_row(
            name="Jonas Nielsen",
            email="jonas.nielsen.import09@skole.dk",
            admission_number="EU-ADM-2026-009",
            class_name="Grade 10",
            section="B",
            roll_number=4,
            gender="Male",
            date_of_birth="2008-02-28",
            phone="+45 20 12 34 56",
            father_name="Henrik Nielsen",
            father_phone="+45 30 23 45 67",
            father_email="henrik.nielsen@mail.dk",
            father_occupation="Wind energy technician",
            father_annual_income=67000,
            mother_name="Mette Nielsen",
            mother_phone="+45 40 34 56 78",
            mother_email="mette.nielsen@mail.dk",
            mother_occupation="HR specialist",
            mother_annual_income=59000,
            current_address="Østerbrogade 55",
            current_city="Copenhagen",
            current_state="Capital Region",
            current_pincode="2100",
            nationality="Danish",
            mother_tongue="Danish",
            place_of_birth="Aarhus",
            blood_group="O+",
            height_cm=176,
            weight_kg=66.0,
            admission_date="2024-06-01",
            student_status="active",
            is_transport_opted=False,
        ),
        build_row(
            name="Zuzanna Nowak",
            email="zuzanna.nowak.import10@edu.pl",
            admission_number="EU-ADM-2026-010",
            class_name="Grade 10",
            section="A",
            roll_number=6,
            gender="Female",
            date_of_birth="2008-10-11",
            phone="+48 533 444 555",
            father_name="Tomasz Nowak",
            father_phone="+48 601 999 888",
            father_email="tomasz.nowak@gmail.com",
            father_occupation="University researcher",
            father_annual_income=54000,
            mother_name="Agnieszka Nowak",
            mother_phone="+48 602 888 777",
            mother_email="agnieszka.nowak@gmail.com",
            mother_occupation="Lab technician",
            mother_annual_income=41000,
            current_address="Floriańska 9",
            current_city="Kraków",
            current_state="Lesser Poland",
            current_pincode="31-019",
            nationality="Polish",
            mother_tongue="Polish",
            place_of_birth="Kraków",
            blood_group="A+",
            height_cm=158,
            weight_kg=49.0,
            admission_date="2024-06-01",
            tc_number="TC-KRK-2024-8891",
            student_status="active",
            is_transport_opted=False,
        ),
    ]

    for r_idx, data in enumerate(rows, start=2):
        for c_idx, val in enumerate(data, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val if val != "" else None)

    out = (
        Path(__file__).resolve().parents[2]
        / "docs"
        / "bulk-import-samples"
        / "bulk-import-european-10-students.xlsx"
    )
    save_wb(wb, out)


def indian_100() -> None:
    random.seed(42)
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    write_headers(ws)

    first_m = [
        "Arjun",
        "Rohan",
        "Vikram",
        "Karan",
        "Aditya",
        "Ishaan",
        "Manav",
        "Dev",
        "Kabir",
        "Siddharth",
    ]
    first_f = [
        "Ananya",
        "Priya",
        "Kavya",
        "Diya",
        "Meera",
        "Sneha",
        "Riya",
        "Ishita",
        "Neha",
        "Pooja",
    ]
    last = [
        "Sharma",
        "Verma",
        "Patel",
        "Reddy",
        "Iyer",
        "Nair",
        "Kapoor",
        "Singh",
        "Mehta",
        "Joshi",
    ]
    cities = [
        ("Ahmedabad", "Gujarat", "380015"),
        ("Jaipur", "Rajasthan", "302001"),
        ("Pune", "Maharashtra", "411004"),
        ("Chennai", "Tamil Nadu", "600004"),
        ("Hyderabad", "Telangana", "500081"),
        ("Kolkata", "West Bengal", "700001"),
        ("Indore", "Madhya Pradesh", "452001"),
        ("Lucknow", "Uttar Pradesh", "226001"),
    ]

    admission_seq = 5000

    def valid_row(i: int, sparse: bool) -> list:
        nonlocal admission_seq
        admission_seq += 1
        is_m = i % 2 == 0
        fn = random.choice(first_m if is_m else first_f)
        ln = random.choice(last)
        name = f"{fn} {ln}"
        email = f"{fn.lower()}.{ln.lower()}.in{admission_seq}@testimport.nexchool.in"
        adm = f"IN-ADM-2026-{admission_seq:04d}"
        section = random.choice(["A", "B"])
        mobile = str(random.randint(6000000000, 9999999999))
        base = {
            "name": name,
            "email": email,
            "admission_number": adm,
            "class_name": "Grade 10",
            "section": section,
            "gender": "Male" if is_m else "Female",
            "date_of_birth": (
                date(2008, 1, 1) + timedelta(days=random.randint(0, 1000))
            ).isoformat(),
            "nationality": "Indian",
            "mother_tongue": random.choice(
                ["Hindi", "Tamil", "Gujarati", "Marathi", ""]
            ),
            "student_status": "active",
            "is_transport_opted": random.choice([True, False, ""]),
        }
        if not sparse:
            city, st, pin = random.choice(cities)
            base.update(
                {
                    "roll_number": (i % 40) + 1,
                    "phone": "+91 " + mobile,
                    "father_name": f"Mr. {random.choice(['Ramesh', 'Suresh', 'Amit'])} {ln}",
                    "father_phone": "+91 " + str(random.randint(7000000000, 9999999999)),
                    "father_email": f"father.{adm.lower()}@gmail.com",
                    "father_occupation": random.choice(
                        [
                            "Business",
                            "Government service",
                            "IT consultant",
                            "Farmer",
                        ]
                    ),
                    "father_annual_income": random.choice(
                        [400000, 850000, 1200000, None]
                    ),
                    "mother_name": f"Mrs. {random.choice(['Sunita', 'Kavita', 'Lakshmi'])} {ln}",
                    "mother_phone": "+91 " + str(random.randint(7000000000, 9999999999)),
                    "current_address": f"{random.randint(1, 120)} {random.choice(['MG Road', 'Ring Road', 'Station Road'])}",
                    "current_city": city,
                    "current_state": st,
                    "current_pincode": pin,
                    "blood_group": random.choice(["A+", "B+", "O+", "AB+"]),
                    "height_cm": random.randint(145, 182),
                    "weight_kg": round(random.uniform(38, 78), 1),
                    "admission_date": "2024-06-01",
                    "religion": random.choice(
                        ["Hindu", "Muslim", "Christian", "Sikh", ""]
                    ),
                    "category": random.choice(
                        ["General", "OBC", "SC", "ST", ""]
                    ),
                }
            )
        else:
            base["roll_number"] = (i % 35) + 1
            base["father_name"] = f"Mr. {random.choice(['Ramesh', 'Suresh'])} {ln}"
            base["father_phone"] = "+91 " + str(random.randint(7000000000, 9999999999))
        return build_row(**base)

    # Scripted edge cases: index i=1..100 maps to Excel row i+1 (row 1 = headers).
    edge_cases: dict[int, dict | None] = {}

    # Row 2: duplicate email with row 5 (set later)
    # Row 5: first occurrence
    # Row 3: invalid email
    edge_cases[3] = {
        "name": "Bad Email Kumar",
        "email": "not-an-email",
        "admission_number": "IN-EDGE-0001",
        "class_name": "Grade 10",
        "section": "A",
    }
    # Row 4: missing section (only class_name)
    edge_cases[4] = {
        "name": "Missing Section Singh",
        "email": "missing.section@testimport.nexchool.in",
        "admission_number": "IN-EDGE-0002",
        "class_name": "Grade 10",
        "section": "",
    }
    # Row 5: valid — will be duplicated email on row 2 — we'll set row 2 after
    edge_cases[5] = {
        "name": "Duplicate Target Reddy",
        "email": "duplicate.email@testimport.nexchool.in",
        "admission_number": "IN-EDGE-0003",
        "class_name": "Grade 10",
        "section": "A",
        "gender": "Male",
        "date_of_birth": "2008-05-15",
    }
    # Row 2: same email as row 5
    edge_cases[2] = {
        "name": "Duplicate Email Person",
        "email": "duplicate.email@testimport.nexchool.in",
        "admission_number": "IN-EDGE-0004",
        "class_name": "Grade 10",
        "section": "B",
    }
    # Row 6: class not found
    edge_cases[6] = {
        "name": "Wrong Class Student",
        "email": "wrong.class@testimport.nexchool.in",
        "admission_number": "IN-EDGE-0005",
        "class_name": "Grade 99",
        "section": "Z",
    }
    # Row 7: invalid date_of_birth
    edge_cases[7] = {
        "name": "Bad Date Naik",
        "email": "bad.date@testimport.nexchool.in",
        "admission_number": "IN-EDGE-0006",
        "class_name": "Grade 10",
        "section": "A",
        "date_of_birth": "31-02-2008",
    }
    # Row 8: duplicate admission with row 9
    edge_cases[8] = {
        "name": "Dup Adm One",
        "email": "dupadm.one@testimport.nexchool.in",
        "admission_number": "IN-DUP-SAME",
        "class_name": "Grade 10",
        "section": "A",
    }
    edge_cases[9] = {
        "name": "Dup Adm Two",
        "email": "dupadm.two@testimport.nexchool.in",
        "admission_number": "IN-DUP-SAME",
        "class_name": "Grade 10",
        "section": "B",
    }
    # Row 10: missing required name
    edge_cases[10] = {
        "name": "",
        "email": "noname@testimport.nexchool.in",
        "admission_number": "IN-EDGE-0007",
        "class_name": "Grade 10",
        "section": "A",
    }
    # Row 11: missing email
    edge_cases[11] = {
        "name": "No Email Student",
        "email": "",
        "admission_number": "IN-EDGE-0008",
        "class_name": "Grade 10",
        "section": "A",
    }
    # Row 12: short phone (invalid soft)
    edge_cases[12] = {
        "name": "Short Phone Das",
        "email": "short.phone@testimport.nexchool.in",
        "admission_number": "IN-EDGE-0009",
        "class_name": "Grade 10",
        "section": "A",
        "phone": "123",
        "father_phone": "45",
    }
    # Row 13: completely blank row (parser skips)
    edge_cases[13] = None
    # Invalid admission_date when provided
    edge_cases[14] = {
        "name": "Bad Admission Date Pillai",
        "email": "bad.admdate@testimport.nexchool.in",
        "admission_number": "IN-EDGE-0010",
        "class_name": "Grade 10",
        "section": "B",
        "admission_date": "not-a-real-date",
    }

    # Fill 1..100 (excel_rows[i] -> Excel row i+1)
    excel_rows: list[list | None] = [None] * 101  # index 0 unused; 1..100
    for i in range(1, 101):
        if i in edge_cases:
            if edge_cases[i] is None:
                excel_rows[i] = [""] * len(HEADERS)  # completely blank
            else:
                excel_rows[i] = build_row(**edge_cases[i])
        else:
            sparse = random.random() < 0.35
            excel_rows[i] = valid_row(i, sparse=sparse)

    for excel_row_num in range(2, 102):
        data = excel_rows[excel_row_num - 1]
        if data is None:
            continue
        for c_idx, val in enumerate(data, start=1):
            v = val
            if v == "":
                v = None
            ws.cell(row=excel_row_num, column=c_idx, value=v)

    out = (
        Path(__file__).resolve().parents[2]
        / "docs"
        / "bulk-import-samples"
        / "bulk-import-indian-100-mixed.xlsx"
    )
    save_wb(wb, out)


def main() -> None:
    european_10()
    indian_100()
    print("Done.")


if __name__ == "__main__":
    main()
