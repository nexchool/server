"""
Excel Import Service

Accepts an .xlsx file with the columns:
    unit_code, programme_code, grade, section, subject?, periods?

Per-row validation: invalid rows are reported in `failed[]` with their
row number; valid rows are inserted, skipping duplicates via the same
structural unique index used by bulk_create_classes. Each row insert is
wrapped in a savepoint so a single bad row never poisons the batch.

If `subject` is provided, the matching active Subject (per tenant) is
attached to the class via class_subjects (uses default weekly_periods=5
or the value in the `periods` column when present).
"""

from __future__ import annotations
from shared.safe_error import safe_error

from io import BytesIO
from typing import Any, BinaryIO, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError

from core.database import db
from modules.academic_programmes.models import AcademicProgramme
from modules.academics.academic_year.models import AcademicYear
from modules.classes.models import Class, ClassSubject
from modules.grades.models import Grade
from modules.school_units.models import SchoolUnit
from modules.subjects.models import Subject


MAX_IMPORT_ROWS = 10_000  # excludes header

REQUIRED_HEADERS = ("unit_code", "programme_code", "grade", "section")
OPTIONAL_HEADERS = ("subject", "periods")

DEFAULT_EXCEL_MAPPING: Dict[str, str] = {
    "unit_code": "unit_code",
    "programme_code": "programme_code",
    "grade": "grade",
    "section": "section",
    "subject": "subject",
    "periods": "periods",
}

_REQUIRED_MAPPING_FIELDS = ("unit_code", "programme_code", "grade", "section")


class UnsupportedFileType(Exception):
    pass


# ---------------------------------------------------------------------------
# Low-level workbook helpers
# ---------------------------------------------------------------------------


def _open_workbook(stream: BinaryIO, filename: str):
    name = (filename or "").lower()
    if not name.endswith(".xlsx"):
        raise UnsupportedFileType(
            f"Unsupported file type: {filename}. Use .xlsx"
        )
    stream.seek(0)
    return load_workbook(stream, read_only=True, data_only=True)


def parse_headers(stream: BinaryIO, filename: str) -> list[str]:
    """Return the first-row cell values as a list of strings."""
    wb = _open_workbook(stream, filename)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        return [str(c).strip() if c is not None else "" for c in row]
    return []


def parse_workbook(
    stream: BinaryIO, filename: str, mapping: Dict[str, str]
) -> list[dict]:
    """
    mapping: {expected_field_name: excel_column_header}
    Returns: list of dicts keyed by expected_field_name.
    Raises UnsupportedFileType for non-.xlsx files.
    Raises ValueError when row count exceeds MAX_IMPORT_ROWS.
    """
    wb = _open_workbook(stream, filename)
    ws = wb.active
    headers: Optional[list[str]] = None
    header_index: Dict[str, int] = {}
    out: list[dict] = []
    count = 0
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(c).strip() if c is not None else "" for c in row]
            header_index = {h: i for i, h in enumerate(headers)}
            continue
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        record: dict = {}
        for field, col in mapping.items():
            idx = header_index.get(col)
            if idx is None or idx >= len(row):
                record[field] = None
            else:
                cell = row[idx]
                record[field] = str(cell).strip() if cell is not None else None
        out.append(record)
        count += 1
        if count > MAX_IMPORT_ROWS:
            raise ValueError(
                f"Import exceeds maximum of {MAX_IMPORT_ROWS} rows. "
                "Split the file and re-import."
            )
    return out


# ---------------------------------------------------------------------------
# Validation helpers
# ---------------------------------------------------------------------------


def _validate_mapping(mapping: Dict[str, str]) -> Optional[str]:
    """Return an error message if the mapping is missing required fields."""
    for field in _REQUIRED_MAPPING_FIELDS:
        if not mapping.get(field, "").strip():
            return f"Mapping must include a non-empty column target for '{field}'"
    return None


def _coerce_periods(value: Any, default: int = 5) -> int:
    if value is None or value == "":
        return default
    try:
        n = int(str(value).strip())
    except (TypeError, ValueError):
        return default
    if n < 1 or n > 40:
        return default
    return n


# ---------------------------------------------------------------------------
# High-level orchestration
# ---------------------------------------------------------------------------


def import_excel(
    tenant_id: str,
    file_storage,
    *,
    academic_year_id: Optional[str],
    mapping: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    """
    Import classes (and optional subject links) from an .xlsx file.

    mapping defaults to DEFAULT_EXCEL_MAPPING.  When a custom mapping is
    supplied (from the column-mapper UI), it must contain non-empty targets
    for all four required fields.

    Returns the same shape as the legacy import_csv:
        {
            success, created, skipped, failed,
            created_count, skipped_count, failed_count,
            subject_links_created, subject_links_skipped,
        }
    """
    if not tenant_id:
        return {"success": False, "error": "Tenant context is required"}
    if not academic_year_id:
        return {"success": False, "error": "academic_year_id is required"}
    if not AcademicYear.query.filter_by(id=academic_year_id, tenant_id=tenant_id).first():
        return {"success": False, "error": "Invalid academic_year_id for this tenant"}

    effective_mapping = mapping if mapping is not None else DEFAULT_EXCEL_MAPPING
    mapping_error = _validate_mapping(effective_mapping)
    if mapping_error:
        return {"success": False, "error": mapping_error}

    try:
        rows = parse_workbook(file_storage.stream, file_storage.filename or "", effective_mapping)
    except UnsupportedFileType as e:
        return {"success": False, "error": str(e)}
    except ValueError as e:
        return {"success": False, "error": str(e)}
    except Exception as e:
        return {"success": False, "error": safe_error(e, "Could not parse file")}

    if not rows:
        return {"success": False, "error": "File is empty or contains only a header row"}

    # Check that the mapped required columns actually exist in the file
    # (parse_workbook sets field to None if the header wasn't found)
    first = rows[0]
    for field in _REQUIRED_MAPPING_FIELDS:
        if field not in first:
            return {"success": False, "error": f"Missing required column mapping: {field}"}

    units = {
        u.code.strip().lower(): u
        for u in SchoolUnit.query.filter_by(tenant_id=tenant_id)
        .filter(SchoolUnit.deleted_at.is_(None))
        .all()
    }
    programmes = {
        p.code.strip().lower(): p
        for p in AcademicProgramme.query.filter_by(tenant_id=tenant_id)
        .filter(AcademicProgramme.deleted_at.is_(None))
        .all()
    }
    grades = {
        g.name.strip().lower(): g
        for g in Grade.query.filter_by(tenant_id=tenant_id)
        .filter(Grade.deleted_at.is_(None))
        .all()
    }
    subjects = {
        s.code.strip().lower(): s
        for s in Subject.query.filter(
            Subject.tenant_id == tenant_id,
            Subject.deleted_at.is_(None),
            Subject.is_active.is_(True),
            Subject.code.isnot(None),
        ).all()
    }
    subjects_by_name = {
        s.name.strip().lower(): s
        for s in Subject.query.filter(
            Subject.tenant_id == tenant_id,
            Subject.deleted_at.is_(None),
            Subject.is_active.is_(True),
        ).all()
    }

    created_rows: List[Dict[str, Any]] = []
    skipped_rows: List[Dict[str, Any]] = []
    error_rows: List[Dict[str, Any]] = []
    subject_links_created = 0
    subject_links_skipped = 0

    for idx, raw in enumerate(rows, start=2):  # row 1 is header
        unit_code = (raw.get("unit_code") or "").strip()
        programme_code = (raw.get("programme_code") or "").strip()
        grade_name = (raw.get("grade") or "").strip()
        section = (raw.get("section") or "").strip()
        subject_token = (raw.get("subject") or "").strip()
        periods = _coerce_periods(raw.get("periods"))

        if not (unit_code and programme_code and grade_name and section):
            error_rows.append({"row_number": idx, "error": "missing required column value"})
            continue

        unit = units.get(unit_code.lower())
        if not unit:
            error_rows.append({"row_number": idx, "error": f"unknown unit_code: {unit_code}"})
            continue
        programme = programmes.get(programme_code.lower())
        if not programme:
            error_rows.append({"row_number": idx, "error": f"unknown programme_code: {programme_code}"})
            continue
        grade = grades.get(grade_name.lower())
        if not grade:
            error_rows.append({"row_number": idx, "error": f"unknown grade: {grade_name}"})
            continue

        existing_class = Class.query.filter_by(
            tenant_id=tenant_id,
            school_unit_id=unit.id,
            programme_id=programme.id,
            grade_id=grade.id,
            section=section,
            academic_year_id=academic_year_id,
        ).first()

        if existing_class:
            class_id = existing_class.id
            skipped_rows.append({"row_number": idx, "class_id": class_id, "reason": "already_exists"})
        else:
            new_cls = Class(
                tenant_id=tenant_id,
                name=f"{grade.name} {section}".strip(),
                section=section,
                academic_year_id=academic_year_id,
                school_unit_id=unit.id,
                programme_id=programme.id,
                grade_id=grade.id,
            )
            try:
                with db.session.begin_nested():
                    db.session.add(new_cls)
                class_id = new_cls.id
                created_rows.append({"row_number": idx, "class_id": class_id})
            except IntegrityError as e:
                msg = str(getattr(e, "orig", e)).lower()
                if "uq_classes_unit_programme_grade_section_year" in msg:
                    existing_class = Class.query.filter_by(
                        tenant_id=tenant_id,
                        school_unit_id=unit.id,
                        programme_id=programme.id,
                        grade_id=grade.id,
                        section=section,
                        academic_year_id=academic_year_id,
                    ).first()
                    class_id = existing_class.id if existing_class else None
                    skipped_rows.append({"row_number": idx, "class_id": class_id, "reason": "already_exists"})
                else:
                    error_rows.append({"row_number": idx, "error": "db_error: constraint violation"})
                    continue

        if subject_token and class_id:
            subject = (
                subjects.get(subject_token.lower())
                or subjects_by_name.get(subject_token.lower())
            )
            if not subject:
                error_rows.append({"row_number": idx, "error": f"unknown subject: {subject_token}"})
                continue
            existing_link = ClassSubject.query.filter(
                ClassSubject.tenant_id == tenant_id,
                ClassSubject.class_id == class_id,
                ClassSubject.subject_id == subject.id,
                ClassSubject.deleted_at.is_(None),
                ClassSubject.status == "active",
            ).first()
            if existing_link:
                subject_links_skipped += 1
            else:
                cs = ClassSubject(
                    tenant_id=tenant_id,
                    class_id=class_id,
                    subject_id=subject.id,
                    weekly_periods=periods,
                )
                try:
                    with db.session.begin_nested():
                        db.session.add(cs)
                    subject_links_created += 1
                except IntegrityError:
                    subject_links_skipped += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return {"success": False, "error": safe_error(e)}

    try:
        from .services import recompute_setup_complete
        recompute_setup_complete(tenant_id)
    except Exception:
        pass

    return {
        "success": True,
        "created": created_rows,
        "skipped": skipped_rows,
        "failed": [
            {**err, "row_number": err.get("row_number", i + 2)}
            for i, err in enumerate(error_rows)
        ],
        "created_count": len(created_rows),
        "skipped_count": len(skipped_rows),
        "failed_count": len(error_rows),
        "subject_links_created": subject_links_created,
        "subject_links_skipped": subject_links_skipped,
    }

