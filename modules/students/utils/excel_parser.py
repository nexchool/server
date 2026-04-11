"""
Parse .xlsx (first sheet, row 1 = headers) using openpyxl.
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, BinaryIO, Dict, List, Tuple, Union

from openpyxl.cell.cell import Cell
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def normalize_header_to_key(header: str) -> str:
    """Lowercase, snake_case: spaces/hyphens -> underscore, strip non-alphanumeric."""
    if header is None:
        return ""
    s = str(header).strip().lower()
    s = re.sub(r"[\s\-]+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _cell_value(cell: Cell) -> Any:
    return cell.value


def _to_serializable(value: Any) -> Any:
    """Normalize cell values for validation (trim strings, date -> ISO)."""
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        return v if v != "" else None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value == int(value):
            return int(value)
        return value
    if isinstance(value, Decimal):
        return float(value)
    return value


def parse_xlsx_to_rows(
    file_obj: Union[BinaryIO, bytes],
) -> Tuple[List[str], List[Dict[str, Any]], List[int]]:
    """
    Read first worksheet. Row 1 = headers (normalized keys).

    Returns:
        header_keys: ordered unique normalized header keys
        rows: dicts per data row; skipped rows are empty (all blanks)
        excel_row_numbers: 1-based Excel row index aligned with rows
    """
    if isinstance(file_obj, bytes):
        stream: BinaryIO = BytesIO(file_obj)
    else:
        file_obj.seek(0)
        stream = file_obj

    wb = load_workbook(stream, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        all_rows = list(ws.iter_rows(values_only=False))
        if not all_rows:
            return [], [], []

        header_cells = all_rows[0]
        raw_headers = [_cell_value(c) for c in header_cells]
        col_to_key = [
            normalize_header_to_key(h) if h is not None else "" for h in raw_headers
        ]

        header_keys: List[str] = []
        seen: set[str] = set()
        for k in col_to_key:
            if k and k not in seen:
                seen.add(k)
                header_keys.append(k)

        data_rows: List[Dict[str, Any]] = []
        row_numbers: List[int] = []

        for r_idx, row_cells in enumerate(all_rows[1:], start=2):
            row_dict: Dict[str, Any] = {}
            any_non_empty = False
            for col_idx, cell in enumerate(row_cells):
                if col_idx >= len(col_to_key):
                    break
                key = col_to_key[col_idx]
                if not key:
                    continue
                raw = _cell_value(cell)
                if raw is not None:
                    if isinstance(raw, str):
                        if raw.strip() != "":
                            any_non_empty = True
                    else:
                        any_non_empty = True
                row_dict[key] = _to_serializable(raw)

            if not any_non_empty:
                continue

            for k in col_to_key:
                if not k:
                    continue
                row_dict.setdefault(k, None)

            data_rows.append(row_dict)
            row_numbers.append(r_idx)
    finally:
        wb.close()

    return header_keys, data_rows, row_numbers
