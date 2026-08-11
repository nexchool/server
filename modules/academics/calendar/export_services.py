"""
Academic Calendar Export

Builds the calendar's export sections once (details, working days, weekly
holidays, public holidays, vacations, semesters, examination windows, school
events) and renders them as CSV, XLSX or PDF. The section data is pulled through
the same services layer that feeds the live calendar (compute_summary + the list
helpers), so an export can never drift from what's published.

Formats:
  csv   — one flat file with a titled block per section (opens in Excel/Sheets)
  excel — a workbook with one sheet per section
  pdf   — a print-ready A4 document (WeasyPrint; degrades to None if unavailable)
"""

from datetime import date, datetime
from io import BytesIO, StringIO
import csv
import re

from flask import g, render_template_string

from modules.academics.academic_year.models import AcademicYear
from modules.academics.backbone.models import AcademicTerm

from . import services
from .holidays import DAY_NAMES, Holiday
from core.school_time import utc_now

try:  # WeasyPrint needs system libs; degrade cleanly when they're missing.
    from weasyprint import HTML
    HAS_WEASYPRINT = True
except (ImportError, OSError):  # pragma: no cover - environment dependent
    HAS_WEASYPRINT = False
    HTML = None  # type: ignore[assignment,misc]

EXPORT_FORMATS = ("csv", "excel", "pdf")

# Section keys, in display order. "overview" is always included; the rest can be
# narrowed via the `sections` argument (maps to the dashboard's kind filters).
ALL_SECTIONS = (
    "weekly_holidays",
    "public_holidays",
    "vacations",
    "semesters",
    "exam_windows",
    "events",
)

_MIME = {
    "csv": "text/csv",
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}
_EXT = {"csv": "csv", "excel": "xlsx", "pdf": "pdf"}


class ExportUnavailableError(RuntimeError):
    """Raised when the requested format cannot be produced (e.g. no WeasyPrint)."""


def _fmt_date(value) -> str:
    if isinstance(value, str):
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return ""


def _weekly_lines(cal) -> list[str]:
    cfg = cal.weekly_config
    lines = [DAY_NAMES[d] for d in sorted(cfg.get("days", []))]
    if cfg.get("second_saturday"):
        lines.append("2nd Saturday")
    if cfg.get("fourth_saturday"):
        lines.append("4th Saturday")
    return lines


def build_export_payload(cal) -> dict:
    """Assemble every section's rows for one calendar (format-agnostic)."""
    year = AcademicYear.query.filter(
        AcademicYear.tenant_id == g.tenant_id,
        AcademicYear.id == cal.academic_year_id,
    ).first()
    summary = services.compute_summary(cal)

    # Public holidays + vacations: dated (non-recurring) rows only; recurring
    # weekly-off rows are already represented in the Weekly Holidays section.
    dated = (
        Holiday.query.filter(
            Holiday.tenant_id == g.tenant_id,
            Holiday.academic_year_id == cal.academic_year_id,
            Holiday.is_recurring.is_(False),
            Holiday.holiday_type != "weekly_off",
            Holiday.start_date.isnot(None),
        )
        .order_by(Holiday.start_date)
        .all()
    )
    public_holidays = [h for h in dated if h.holiday_type != "vacation"]
    vacations = [h for h in dated if h.holiday_type == "vacation"]

    terms = (
        AcademicTerm.query.filter(
            AcademicTerm.tenant_id == g.tenant_id,
            AcademicTerm.academic_year_id == cal.academic_year_id,
            AcademicTerm.deleted_at.is_(None),
        )
        .order_by(AcademicTerm.sequence)
        .all()
    )
    exams = services.list_exam_windows(cal.academic_year_id, active_only=True)
    events = services.list_school_events(cal.academic_year_id, active_only=True)

    year_label = year.name if year else cal.academic_year_id

    return {
        "title": f"Academic Calendar — {year_label}",
        "year_label": year_label,
        "generated_at": utc_now().strftime("%d %b %Y, %H:%M UTC"),
        "status": cal.status,
        "summary": summary,
        "weekly_holidays": _weekly_lines(cal),
        "public_holidays": [
            {
                "name": h.name,
                "type": h.holiday_type,
                "start_date": _fmt_date(h.start_date),
                "end_date": _fmt_date(h.end_date or h.start_date),
                "applies_to": getattr(h, "applies_to", "entire_school"),
                "description": h.description or "",
            }
            for h in public_holidays
        ],
        "vacations": [
            {
                "name": h.name,
                "start_date": _fmt_date(h.start_date),
                "end_date": _fmt_date(h.end_date or h.start_date),
                "description": h.description or "",
            }
            for h in vacations
        ],
        "semesters": [
            {
                "name": t.name,
                "start_date": _fmt_date(t.start_date),
                "end_date": _fmt_date(t.end_date),
            }
            for t in terms
        ],
        "exam_windows": [
            {
                "name": w.name,
                "type": w.exam_type,
                "start_date": _fmt_date(w.start_date),
                "end_date": _fmt_date(w.end_date),
                "status": w.status,
                "description": w.description or "",
            }
            for w in exams
        ],
        "events": [
            {
                "name": e.name,
                "type": e.event_type,
                "date": _fmt_date(e.event_date),
                "applies_to": e.applies_to,
                "status": e.status,
                "description": e.description or "",
            }
            for e in events
        ],
    }


def _resolve_sections(sections) -> tuple[str, ...]:
    """Normalise the requested section list; default to everything."""
    if not sections:
        return ALL_SECTIONS
    wanted = {s.strip() for s in sections if s and s.strip() in ALL_SECTIONS}
    return tuple(s for s in ALL_SECTIONS if s in wanted) or ALL_SECTIONS


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (value or "calendar").lower()).strip("-") or "calendar"


# ---------------------------------------------------------------------------
# Renderers
# ---------------------------------------------------------------------------

def _to_csv(payload: dict, sections: tuple[str, ...]) -> bytes:
    buf = StringIO()
    w = csv.writer(buf)
    s = payload["summary"]

    w.writerow([payload["title"]])
    w.writerow(["Status", payload["status"]])
    w.writerow(["Generated", payload["generated_at"]])
    w.writerow([])
    w.writerow(["Overview"])
    w.writerow(["Total Days", s["total_days"]])
    w.writerow(["Working Days", s["working_days"]])
    w.writerow(["Weekly Holiday Days", s["weekly_holiday_days"]])
    w.writerow(["Public Holiday Days", s["public_holiday_days"]])
    w.writerow(["Vacation Days", s["vacation_days"]])
    w.writerow(["Exam Days", s["exam_days"]])
    w.writerow(["Semesters", s["semester_count"]])
    w.writerow(["Exam Windows", s["exam_window_count"]])
    w.writerow(["Events", s["event_count"]])

    if "weekly_holidays" in sections:
        w.writerow([])
        w.writerow(["Weekly Holidays"])
        for line in payload["weekly_holidays"] or ["—"]:
            w.writerow([line])

    if "public_holidays" in sections:
        w.writerow([])
        w.writerow(["Public Holidays"])
        w.writerow(["Name", "Type", "Start Date", "End Date", "Applies To", "Description"])
        for h in payload["public_holidays"]:
            w.writerow([h["name"], h["type"], h["start_date"], h["end_date"], h["applies_to"], h["description"]])

    if "vacations" in sections:
        w.writerow([])
        w.writerow(["Vacations"])
        w.writerow(["Name", "Start Date", "End Date", "Description"])
        for v in payload["vacations"]:
            w.writerow([v["name"], v["start_date"], v["end_date"], v["description"]])

    if "semesters" in sections:
        w.writerow([])
        w.writerow(["Semesters"])
        w.writerow(["Name", "Start Date", "End Date"])
        for t in payload["semesters"]:
            w.writerow([t["name"], t["start_date"], t["end_date"]])

    if "exam_windows" in sections:
        w.writerow([])
        w.writerow(["Examination Windows"])
        w.writerow(["Name", "Type", "Start Date", "End Date", "Status", "Description"])
        for x in payload["exam_windows"]:
            w.writerow([x["name"], x["type"], x["start_date"], x["end_date"], x["status"], x["description"]])

    if "events" in sections:
        w.writerow([])
        w.writerow(["School Events"])
        w.writerow(["Name", "Type", "Date", "Applies To", "Status", "Description"])
        for e in payload["events"]:
            w.writerow([e["name"], e["type"], e["date"], e["applies_to"], e["status"], e["description"]])

    return buf.getvalue().encode("utf-8-sig")


def _autosize(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 60)


def _to_excel(payload: dict, sections: tuple[str, ...]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    s = payload["summary"]

    ws = wb.active
    ws.title = "Overview"
    ws.append([payload["title"]])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Status", payload["status"]])
    ws.append(["Generated", payload["generated_at"]])
    ws.append([])
    for label, key in (
        ("Total Days", "total_days"), ("Working Days", "working_days"),
        ("Weekly Holiday Days", "weekly_holiday_days"),
        ("Public Holiday Days", "public_holiday_days"),
        ("Vacation Days", "vacation_days"), ("Exam Days", "exam_days"),
        ("Semesters", "semester_count"), ("Exam Windows", "exam_window_count"),
        ("Events", "event_count"),
    ):
        ws.append([label, s[key]])
    _autosize(ws)

    def sheet(title: str, header: list[str], rows: list[list]) -> None:
        w = wb.create_sheet(title=title[:31])
        w.append(header)
        for c in w[1]:
            c.font = Font(bold=True)
        for r in rows:
            w.append(r)
        _autosize(w)

    if "weekly_holidays" in sections:
        sheet("Weekly Holidays", ["Weekly Holiday"], [[x] for x in payload["weekly_holidays"]])
    if "public_holidays" in sections:
        sheet("Public Holidays", ["Name", "Type", "Start Date", "End Date", "Applies To", "Description"],
              [[h["name"], h["type"], h["start_date"], h["end_date"], h["applies_to"], h["description"]]
               for h in payload["public_holidays"]])
    if "vacations" in sections:
        sheet("Vacations", ["Name", "Start Date", "End Date", "Description"],
              [[v["name"], v["start_date"], v["end_date"], v["description"]] for v in payload["vacations"]])
    if "semesters" in sections:
        sheet("Semesters", ["Name", "Start Date", "End Date"],
              [[t["name"], t["start_date"], t["end_date"]] for t in payload["semesters"]])
    if "exam_windows" in sections:
        sheet("Exam Windows", ["Name", "Type", "Start Date", "End Date", "Status", "Description"],
              [[x["name"], x["type"], x["start_date"], x["end_date"], x["status"], x["description"]]
               for x in payload["exam_windows"]])
    if "events" in sections:
        sheet("School Events", ["Name", "Type", "Date", "Applies To", "Status", "Description"],
              [[e["name"], e["type"], e["date"], e["applies_to"], e["status"], e["description"]]
               for e in payload["events"]])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


_PDF_TEMPLATE = """
<!doctype html><html><head><meta charset="utf-8"><style>
  @page { size: A4; margin: 1.6cm; }
  body { font-family: 'Helvetica Neue', Arial, sans-serif; color: #1f2937; font-size: 11px; }
  h1 { font-size: 20px; margin: 0 0 2px; }
  .meta { color: #6b7280; font-size: 11px; margin-bottom: 16px; }
  h2 { font-size: 13px; margin: 18px 0 6px; padding-bottom: 3px; border-bottom: 2px solid #e5e7eb; }
  table { width: 100%; border-collapse: collapse; margin-top: 4px; }
  th, td { text-align: left; padding: 5px 8px; border-bottom: 1px solid #f0f0f0; vertical-align: top; }
  th { background: #f9fafb; font-size: 10px; text-transform: uppercase; letter-spacing: .03em; color: #6b7280; }
  .stats { display: flex; flex-wrap: wrap; gap: 8px; margin-top: 6px; }
  .stat { border: 1px solid #e5e7eb; border-radius: 6px; padding: 6px 12px; min-width: 90px; }
  .stat .n { font-size: 16px; font-weight: 700; }
  .stat .l { font-size: 10px; color: #6b7280; }
  .chips span { display: inline-block; border: 1px solid #e5e7eb; border-radius: 999px; padding: 2px 10px; margin: 2px 4px 2px 0; }
  .empty { color: #9ca3af; font-style: italic; }
</style></head><body>
  <h1>{{ payload.title }}</h1>
  <div class="meta">Status: {{ payload.status|capitalize }} · Generated {{ payload.generated_at }}</div>

  <h2>Overview</h2>
  <div class="stats">
    {% for l, n in overview %}<div class="stat"><div class="n">{{ n }}</div><div class="l">{{ l }}</div></div>{% endfor %}
  </div>

  {% if 'weekly_holidays' in sections %}
  <h2>Weekly Holidays</h2>
  {% if payload.weekly_holidays %}<div class="chips">{% for w in payload.weekly_holidays %}<span>{{ w }}</span>{% endfor %}</div>
  {% else %}<p class="empty">None configured.</p>{% endif %}
  {% endif %}

  {% if 'public_holidays' in sections %}
  <h2>Public Holidays</h2>
  {% if payload.public_holidays %}<table><thead><tr><th>Name</th><th>Type</th><th>Date(s)</th><th>Applies To</th></tr></thead><tbody>
  {% for h in payload.public_holidays %}<tr><td>{{ h.name }}</td><td>{{ h.type }}</td><td>{{ h.start_date }}{% if h.end_date != h.start_date %} – {{ h.end_date }}{% endif %}</td><td>{{ h.applies_to }}</td></tr>{% endfor %}
  </tbody></table>{% else %}<p class="empty">None.</p>{% endif %}
  {% endif %}

  {% if 'vacations' in sections %}
  <h2>Vacations</h2>
  {% if payload.vacations %}<table><thead><tr><th>Name</th><th>Start</th><th>End</th></tr></thead><tbody>
  {% for v in payload.vacations %}<tr><td>{{ v.name }}</td><td>{{ v.start_date }}</td><td>{{ v.end_date }}</td></tr>{% endfor %}
  </tbody></table>{% else %}<p class="empty">None.</p>{% endif %}
  {% endif %}

  {% if 'semesters' in sections %}
  <h2>Semesters</h2>
  {% if payload.semesters %}<table><thead><tr><th>Name</th><th>Start</th><th>End</th></tr></thead><tbody>
  {% for t in payload.semesters %}<tr><td>{{ t.name }}</td><td>{{ t.start_date }}</td><td>{{ t.end_date }}</td></tr>{% endfor %}
  </tbody></table>{% else %}<p class="empty">None.</p>{% endif %}
  {% endif %}

  {% if 'exam_windows' in sections %}
  <h2>Examination Windows</h2>
  {% if payload.exam_windows %}<table><thead><tr><th>Name</th><th>Type</th><th>Start</th><th>End</th></tr></thead><tbody>
  {% for x in payload.exam_windows %}<tr><td>{{ x.name }}</td><td>{{ x.type }}</td><td>{{ x.start_date }}</td><td>{{ x.end_date }}</td></tr>{% endfor %}
  </tbody></table>{% else %}<p class="empty">None.</p>{% endif %}
  {% endif %}

  {% if 'events' in sections %}
  <h2>School Events</h2>
  {% if payload.events %}<table><thead><tr><th>Name</th><th>Type</th><th>Date</th><th>Applies To</th></tr></thead><tbody>
  {% for e in payload.events %}<tr><td>{{ e.name }}</td><td>{{ e.type }}</td><td>{{ e.date }}</td><td>{{ e.applies_to }}</td></tr>{% endfor %}
  </tbody></table>{% else %}<p class="empty">None.</p>{% endif %}
  {% endif %}
</body></html>
"""


def _to_pdf(payload: dict, sections: tuple[str, ...]) -> bytes:
    if not HAS_WEASYPRINT or HTML is None:
        raise ExportUnavailableError("PDF export is unavailable on this server.")
    s = payload["summary"]
    overview = [
        ("Total Days", s["total_days"]), ("Working Days", s["working_days"]),
        ("Weekly Holidays", s["weekly_holiday_days"]),
        ("Public Holidays", s["public_holiday_days"]),
        ("Vacation Days", s["vacation_days"]), ("Exam Days", s["exam_days"]),
        ("Semesters", s["semester_count"]), ("Events", s["event_count"]),
    ]
    html = render_template_string(
        _PDF_TEMPLATE, payload=payload, sections=list(sections), overview=overview
    )
    return HTML(string=html).write_pdf()


def export_calendar(cal, fmt: str, sections=None) -> tuple[bytes, str, str]:
    """Render one calendar as (content_bytes, mimetype, download_filename)."""
    fmt = (fmt or "").strip().lower()
    if fmt not in EXPORT_FORMATS:
        raise ValueError(f"Unsupported format '{fmt}'. Use one of: {', '.join(EXPORT_FORMATS)}")

    resolved = _resolve_sections(sections)
    payload = build_export_payload(cal)
    if fmt == "csv":
        content = _to_csv(payload, resolved)
    elif fmt == "excel":
        content = _to_excel(payload, resolved)
    else:
        content = _to_pdf(payload, resolved)

    stamp = utc_now().strftime("%Y%m%d")
    filename = f"academic-calendar-{_slug(payload['year_label'])}-{stamp}.{_EXT[fmt]}"
    return content, _MIME[fmt], filename
